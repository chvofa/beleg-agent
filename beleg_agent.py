#!/usr/bin/env python3
"""
Beleg-Agent – Automatische Verarbeitung von Rechnungen/Belegen.

Überwacht die _Inbox auf neue PDF/Bild-Dateien, extrahiert Rechnungsdaten
via Claude Vision API und legt die Belege strukturiert ab.
"""

import base64
import io
import json
import logging
import os
import re
import shutil
import sys
import time
from datetime import datetime, date
from pathlib import Path

import anthropic
import pypdfium2 as pdfium
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

import config

# ── Toast Icon ────────────────────────────────────────────────────────────
_TOAST_ICON = os.path.join(os.path.dirname(os.path.abspath(__file__)), "beleg-agent-icon.png")

def _erstelle_toast_icon():
    """Erstellt ein Icon fuer Toast-Benachrichtigungen."""
    if os.path.exists(_TOAST_ICON):
        return
    try:
        from PIL import Image, ImageDraw, ImageFont
        size = 128
        img = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        margin = size // 16
        draw.ellipse(
            [margin, margin, size - margin, size - margin],
            fill=(46, 160, 67),
            outline=(255, 255, 255, 200),
            width=max(2, size // 32),
        )
        try:
            font = ImageFont.truetype("segoeuib.ttf", int(size * 0.5))
        except Exception:
            try:
                font = ImageFont.truetype("arial.ttf", int(size * 0.5))
            except Exception:
                font = ImageFont.load_default()
        bbox = draw.textbbox((0, 0), "B", font=font)
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        x = (size - tw) // 2
        y = (size - th) // 2 - bbox[1]
        draw.text((x, y), "B", fill=(255, 255, 255), font=font)
        img.save(_TOAST_ICON, format="PNG")
    except Exception:
        pass

_erstelle_toast_icon()

# ── Windows Toast Notifications ────────────────────────────────────────────
try:
    from winotify import Notification

    def toast(title: str, msg: str):
        """Sendet Windows Toast-Benachrichtigung."""
        try:
            t = Notification(app_id="Beleg-Agent", title=title, msg=msg, duration="short")
            if os.path.exists(_TOAST_ICON):
                t.set_audio(None, suppress=True)
                t.icon = _TOAST_ICON
            t.show()
        except Exception:
            pass  # Notification-Fehler sollen Agent nicht stoppen
except ImportError:
    def toast(title: str, msg: str):
        pass  # winotify nicht installiert, still ignorieren


# ── Windows UTF-8 Konsole ─────────────────────────────────────────────────
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Logging ────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(config.LOG_DATEI, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("beleg-agent")


# ── Claude API Client ──────────────────────────────────────────────────────
client = anthropic.Anthropic()  # liest ANTHROPIC_API_KEY aus Umgebung


# ═══════════════════════════════════════════════════════════════════════════
#  Hilfsfunktionen
# ═══════════════════════════════════════════════════════════════════════════

MAX_PDF_SEITEN = 5  # Maximal so viele Seiten an die API senden


def datei_zu_bilder_base64(dateipfad: str) -> list[tuple[str, str]]:
    """Konvertiert PDF (alle Seiten bis MAX) oder Bild → Liste von (base64, media_type)."""
    endung = Path(dateipfad).suffix.lower()

    if endung == ".pdf":
        pdf = pdfium.PdfDocument(dateipfad)
        anzahl = min(len(pdf), MAX_PDF_SEITEN)
        bilder = []
        for i in range(anzahl):
            bitmap = pdf[i].render(scale=2)
            pil_image = bitmap.to_pil()
            buf = io.BytesIO()
            pil_image.save(buf, format="PNG")
            bilder.append((base64.standard_b64encode(buf.getvalue()).decode("utf-8"), "image/png"))
        pdf.close()
        return bilder

    # Bild-Dateien
    with Image.open(dateipfad) as img:
        if img.mode == "RGBA":
            img = img.convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return [(base64.standard_b64encode(buf.getvalue()).decode("utf-8"), "image/png")]


def extrahiere_rechnungsdaten(dateipfad: str) -> dict | None:
    """Sendet alle Seiten an Claude Vision API und extrahiert strukturierte Rechnungsdaten."""
    try:
        bilder = datei_zu_bilder_base64(dateipfad)
    except Exception as e:
        log.error(f"Fehler bei Bildkonvertierung von {dateipfad}: {e}")
        return None

    log.info(f"Sende {len(bilder)} Seite(n) an Claude API...")

    prompt = """Analysiere alle Seiten dieses Dokuments (Rechnung / Beleg) und extrahiere folgende Daten.
Antworte ausschließlich mit validem JSON, kein anderer Text.

{
  "rechnungssteller": "Name des Lieferanten/Rechnungsstellers",
  "rechnungsdatum": "YYYY-MM-DD",
  "betrag": 123.45,
  "waehrung": "CHF",
  "typ": "Rechnung",
  "zahlungsart": "",
  "ist_paypal": false,
  "confidence": {
    "rechnungssteller": 0.95,
    "rechnungsdatum": 0.90,
    "betrag": 0.95,
    "waehrung": 0.95,
    "typ": 0.95,
    "zahlungsart": 0.50
  }
}

Regeln:
- rechnungssteller: Kurzname des Unternehmens (z.B. "Amazon", "Migros", "Swisscom")
- rechnungsdatum: Falls nicht erkennbar, verwende das heutige Datum und setze confidence niedrig
- betrag: Gesamtbetrag / Endbetrag / Saldo als Zahl (kein Tausendertrennzeichen), immer positiv.
  Kann auf einer späteren Seite stehen (z.B. "Saldo zu Ihren Gunsten" bei Versicherungen).
- waehrung: ISO-Code der Rechnungswährung (CHF, EUR, USD, GBP, etc.)
- typ: "Rechnung" (wir zahlen) oder "Gutschrift" (wir erhalten Geld zurück).
  Gutschrift = Rückerstattung, Credit Note, Rückvergütung, Versicherungsleistung, Kundenzahlung an uns.
  Im Zweifel: "Rechnung".
- zahlungsart: Wie wurde bezahlt? Mögliche Werte:
    "KK CHF" = Kreditkarte mit CHF-Abrechnung (z.B. Firmen-Visa CHF)
    "KK EUR" = Kreditkarte mit EUR-Abrechnung
    "Überweisung" = Banküberweisung / Einzahlung / LSV
    "" = leer lassen wenn Zahlungsart nicht erkennbar
  Hinweis: Die Rechnungswährung (z.B. USD) kann sich von der Karte unterscheiden.
  Achte auf Hinweise wie Kartennummer, Firmenname auf KK, Zahlungsmethode, VISA, etc.
- ist_paypal: true wenn PayPal als Zahlungsweg erkennbar ist (PayPal-Logo, PayPal-Transaktionscode, etc.)
  Hinweis: PayPal und Kreditkarte schliessen sich nicht aus – PayPal kann über KK laufen.
- confidence: Dein Vertrauenswert 0.0-1.0 für jedes Feld.
  Setze zahlungsart-confidence auf 0.0 wenn du keine Zahlungsinfo findest (Feld bleibt leer).
"""

    try:
        # Alle Seiten als Bild-Blöcke senden
        content_blocks = []
        for bild_b64, media_type in bilder:
            content_blocks.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": bild_b64,
                },
            })
        content_blocks.append({"type": "text", "text": prompt})

        response = client.messages.create(
            model=config.ANTHROPIC_MODEL,
            max_tokens=1024,
            messages=[{"role": "user", "content": content_blocks}],
        )
    except anthropic.APIError as e:
        log.error(f"Claude API-Fehler: {e}")
        return None

    # JSON aus Antwort parsen
    antwort_text = response.content[0].text.strip()
    # Falls Claude Markdown-Codeblock nutzt
    json_match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", antwort_text, re.DOTALL)
    if json_match:
        antwort_text = json_match.group(1)

    try:
        daten = json.loads(antwort_text)
    except json.JSONDecodeError:
        log.error(f"Konnte JSON nicht parsen: {antwort_text[:200]}")
        return None

    # Gesamt-Confidence berechnen
    # zahlungsart-Confidence wird nur einbezogen wenn Zahlungsart erkannt wurde,
    # damit eine fehlende Zahlungsinfo den Score nicht drückt.
    conf = daten.get("confidence", {})
    if conf:
        kernfelder = ["rechnungssteller", "rechnungsdatum", "betrag", "waehrung"]
        werte = [conf[f] for f in kernfelder if f in conf]
        # Zahlungsart nur einbeziehen wenn erkannt (confidence > 0)
        za_conf = conf.get("zahlungsart", 0)
        if za_conf > 0:
            werte.append(za_conf)
        daten["gesamt_confidence"] = round(sum(werte) / len(werte), 3) if werte else 0.0
    else:
        daten["gesamt_confidence"] = 0.0

    # Zahlungsart aus Protokoll-Historie ergänzen wenn unbekannt
    if not daten.get("zahlungsart"):
        za_hist, pp_hist = lookup_zahlungsart(daten.get("rechnungssteller", ""))
        if za_hist:
            daten["zahlungsart"] = za_hist
            log.info(f"Zahlungsart aus Historie uebernommen: {za_hist} (gleicher Rechnungssteller)")
        if pp_hist is not None and not daten.get("ist_paypal"):
            daten["ist_paypal"] = pp_hist

    return daten


def lookup_zahlungsart(rechnungssteller: str) -> tuple[str, bool | None]:
    """Sucht im Protokoll nach bekannter Zahlungsart fuer diesen Rechnungssteller."""
    if not rechnungssteller or not os.path.exists(config.EXCEL_PROTOKOLL):
        return "", None

    try:
        wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL, read_only=True)
        ws = wb.active
    except Exception:
        return "", None

    rs_neu = rechnungssteller.lower().strip()
    zahlungsart = ""
    ist_paypal = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        rs_alt = str(row[1] or "").lower().strip()
        # Fuzzy-Match: Teilstring in beide Richtungen
        if rs_neu in rs_alt or rs_alt in rs_neu:
            za = str(row[5] or "").strip()
            pp = str(row[6] or "").strip()
            if za and za in ("KK CHF", "KK EUR", "Überweisung"):
                zahlungsart = za
            if pp == "Ja":
                ist_paypal = True

    wb.close()
    return zahlungsart, ist_paypal


def baue_zieldateiname(daten: dict, original_endung: str) -> str:
    """Baut Dateinamen: [Gutschrift ][Rechnungssteller] - [Währung] [Betrag][ KK CHF/EUR].ext"""
    rs = daten["rechnungssteller"].strip()
    # Ungültige Dateinamen-Zeichen entfernen
    rs = re.sub(r'[<>:"/\\|?*]', "", rs)
    waehrung = daten["waehrung"].upper()
    betrag = f"{daten['betrag']:.2f}"

    # Gutschrift-Prefix
    prefix = "Gutschrift " if daten.get("typ") == "Gutschrift" else ""

    # Zahlungsart-Suffix: nur bei bekannter KK
    zahlungsart = daten.get("zahlungsart", "")
    if zahlungsart in ("KK CHF", "KK EUR"):
        suffix = f" {zahlungsart}"
    else:
        suffix = ""

    return f"{prefix}{rs} - {waehrung} {betrag}{suffix}{original_endung}"


def baue_zielordner(daten: dict) -> str:
    """Baut den Zielordner basierend auf dem Rechnungsdatum."""
    try:
        datum = datetime.strptime(daten["rechnungsdatum"], "%Y-%m-%d").date()
    except (ValueError, KeyError):
        datum = date.today()

    jahr = str(datum.year)
    monat_ordner = config.MONATE.get(datum.month, f"{datum.month:02d}_Unbekannt")
    return os.path.join(config.ABLAGE_STAMMPFAD, jahr, monat_ordner)


def pruefe_duplikat_protokoll(daten: dict) -> bool:
    """Prüft ob ein ähnlicher Eintrag bereits im Excel-Protokoll existiert (fuzzy)."""
    if not os.path.exists(config.EXCEL_PROTOKOLL):
        return False

    try:
        wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
        ws = wb.active
    except Exception:
        return False

    rs_neu = daten.get("rechnungssteller", "").lower().strip()
    betrag_neu = daten.get("betrag", 0)
    datum_neu = daten.get("rechnungsdatum", "")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        # Spalten: Datum(0), Rechnungssteller(1), Betrag(2)
        datum_alt = str(row[0]).strip() if row[0] else ""
        rs_alt = str(row[1]).lower().strip() if row[1] else ""
        try:
            betrag_alt = float(row[2]) if row[2] else 0
        except (ValueError, TypeError):
            betrag_alt = 0

        # Fuzzy: gleicher Steller (Teilstring), gleicher Betrag, gleiches Datum
        if (
            (rs_neu in rs_alt or rs_alt in rs_neu)
            and abs(betrag_alt - betrag_neu) < 0.01
            and datum_neu == datum_alt
        ):
            wb.close()
            return True

    wb.close()
    return False


def pruefe_duplikat_datei(zielordner: str, dateiname: str) -> bool:
    """Prüft ob eine Datei mit gleichem Namen bereits im Zielordner existiert."""
    return os.path.exists(os.path.join(zielordner, dateiname))


def erstelle_excel_wenn_noetig():
    """Erstellt das Excel-Protokoll mit Header falls es noch nicht existiert."""
    if os.path.exists(config.EXCEL_PROTOKOLL):
        return

    os.makedirs(os.path.dirname(config.EXCEL_PROTOKOLL), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Belege"

    # Header-Zeile
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for col_idx, spalte in enumerate(config.EXCEL_SPALTEN, start=1):
        zelle = ws.cell(row=1, column=col_idx, value=spalte)
        zelle.font = header_font
        zelle.fill = header_fill
        zelle.alignment = Alignment(horizontal="center")

    # Spaltenbreiten
    breiten = [15, 25, 12, 10, 12, 14, 10, 35, 60, 12, 15, 18, 40]
    for i, b in enumerate(breiten, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = b

    wb.save(config.EXCEL_PROTOKOLL)
    wb.close()
    log.info(f"Excel-Protokoll erstellt: {config.EXCEL_PROTOKOLL}")


def schreibe_protokoll(daten: dict, original_name: str, ablagepfad: str):
    """Fügt einen neuen Eintrag ins Excel-Protokoll ein."""
    erstelle_excel_wenn_noetig()

    try:
        wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
        ws = wb.active

        neue_zeile = [
            daten.get("rechnungsdatum", ""),
            daten.get("rechnungssteller", ""),
            daten.get("betrag", 0),
            daten.get("waehrung", ""),
            daten.get("typ", "Rechnung"),           # Rechnung / Gutschrift
            daten.get("zahlungsart", ""),            # KK CHF / KK EUR / Überweisung / leer
            "Ja" if daten.get("ist_paypal") else "Nein",
            original_name,
            ablagepfad,
            "Nein",
            daten.get("gesamt_confidence", 0),
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            "",                                      # Bemerkungen
        ]
        ws.append(neue_zeile)
        wb.save(config.EXCEL_PROTOKOLL)
        wb.close()
        log.info("Protokoll-Eintrag geschrieben.")
    except Exception as e:
        log.error(f"Fehler beim Schreiben ins Excel-Protokoll: {e}")
        log.error("Ablage wird trotzdem durchgeführt.")


def terminal_rueckfrage(daten: dict, dateiname: str) -> dict | None:
    """Zeigt extrahierte Daten im Terminal an und fragt den Benutzer."""
    print("\n" + "=" * 60)
    print(f"  📄 RÜCKFRAGE: {dateiname}")
    print("=" * 60)
    zahlungsart = daten.get('zahlungsart', '') or '(unbekannt)'
    typ = daten.get('typ', 'Rechnung')
    print(f"  Typ              : {typ}")
    print(f"  Rechnungssteller : {daten.get('rechnungssteller', '?')}")
    print(f"  Rechnungsdatum   : {daten.get('rechnungsdatum', '?')}")
    print(f"  Betrag           : {daten.get('betrag', '?')}")
    print(f"  Währung          : {daten.get('waehrung', '?')}")
    print(f"  Zahlungsart      : {zahlungsart}")
    print(f"  PayPal           : {'Ja' if daten.get('ist_paypal') else 'Nein'}")
    print(f"  Confidence       : {daten.get('gesamt_confidence', 0):.1%}")
    print("-" * 60)
    print("  [Enter] = Übernehmen")
    print("  [a]     = Ablehnen (→ [PRÜFEN]-Prefix)")
    print("  Oder Korrekturen eingeben als JSON-Felder, z.B.:")
    print('  {"rechnungssteller": "Swisscom", "betrag": 99.90}')
    print(f"  (Timeout: {config.RÜCKFRAGE_TIMEOUT_SEKUNDEN}s → automatisch [PRÜFEN])")
    print("=" * 60)

    # Timeout-basierte Eingabe
    import select
    if sys.platform == "win32":
        # Windows: Thread-basierter Timeout
        import threading
        ergebnis = {"eingabe": None}

        def lese_eingabe():
            try:
                ergebnis["eingabe"] = input("  > ").strip()
            except EOFError:
                ergebnis["eingabe"] = "a"

        thread = threading.Thread(target=lese_eingabe, daemon=True)
        thread.start()
        thread.join(timeout=config.RÜCKFRAGE_TIMEOUT_SEKUNDEN)

        if thread.is_alive():
            print("\n  ⏰ Timeout – Datei wird mit [PRÜFEN] markiert.")
            return None

        eingabe = ergebnis["eingabe"]
    else:
        # Unix
        ready, _, _ = select.select([sys.stdin], [], [], config.RÜCKFRAGE_TIMEOUT_SEKUNDEN)
        if not ready:
            print("\n  ⏰ Timeout – Datei wird mit [PRÜFEN] markiert.")
            return None
        eingabe = sys.stdin.readline().strip()

    if not eingabe:
        # Enter → übernehmen wie erkannt
        return daten

    if eingabe.lower() == "a":
        return None

    # Versuche JSON-Korrekturen zu parsen
    try:
        korrekturen = json.loads(eingabe)
        daten.update(korrekturen)
        daten["gesamt_confidence"] = 1.0  # Manuell bestätigt
        print("  ✅ Korrekturen übernommen.")
        return daten
    except json.JSONDecodeError:
        print("  ⚠️ Ungültige Eingabe – Datei wird mit [PRÜFEN] markiert.")
        return None


def markiere_als_pruefen(dateipfad: str) -> str:
    """Benennt die Datei mit [PRÜFEN]-Prefix um."""
    ordner = os.path.dirname(dateipfad)
    name = os.path.basename(dateipfad)

    if name.startswith("[PRÜFEN]_"):
        return dateipfad

    neuer_name = f"[PRÜFEN]_{name}"
    neuer_pfad = os.path.join(ordner, neuer_name)
    os.rename(dateipfad, neuer_pfad)
    log.info(f"Markiert als [PRÜFEN]: {neuer_name}")
    toast("Beleg prüfen", f"{name} - bitte manuell prüfen in _Inbox")
    return neuer_pfad


def lege_datei_ab(dateipfad: str, daten: dict) -> bool:
    """Verschiebt die Datei in den korrekten Ablageordner und schreibt Protokoll."""
    original_name = os.path.basename(dateipfad)
    endung = Path(dateipfad).suffix.lower()

    # Zielordner und Dateiname bauen
    zielordner = baue_zielordner(daten)
    neuer_name = baue_zieldateiname(daten, endung)

    # Duplikat-Checks
    if pruefe_duplikat_datei(zielordner, neuer_name):
        log.warning(f"Duplikat (Dateiname): {neuer_name} existiert bereits in {zielordner}")
        # Suffix anhängen
        stamm = Path(neuer_name).stem
        neuer_name = f"{stamm} (2){endung}"

    if pruefe_duplikat_protokoll(daten):
        log.warning(
            f"Mögliches Duplikat im Protokoll: "
            f"{daten['rechnungssteller']} / {daten['betrag']} / {daten['rechnungsdatum']}"
        )
        toast("Duplikat erkannt", f"{daten['rechnungssteller']} {daten['waehrung']} {daten['betrag']:.2f} - als [PRÜFEN] markiert")
        markiere_als_pruefen(dateipfad)
        return False

    # Ordner erstellen und Datei verschieben
    os.makedirs(zielordner, exist_ok=True)
    zielpfad = os.path.join(zielordner, neuer_name)

    try:
        shutil.move(dateipfad, zielpfad)
        log.info(f"Abgelegt: {original_name} → {zielpfad}")
    except Exception as e:
        log.error(f"Fehler beim Verschieben: {e}")
        toast("Fehler", f"Konnte {original_name} nicht verschieben")
        return False

    # Protokoll schreiben (Fehler hier soll Ablage nicht rückgängig machen)
    schreibe_protokoll(daten, original_name, zielpfad)

    # Toast-Benachrichtigung
    typ = daten.get("typ", "Rechnung")
    rs = daten.get("rechnungssteller", "?")
    betrag = daten.get("betrag", 0)
    waehrung = daten.get("waehrung", "")
    toast(f"Beleg abgelegt", f"{typ}: {rs} - {waehrung} {betrag:.2f}")

    return True


# ═══════════════════════════════════════════════════════════════════════════
#  Verarbeitungs-Pipeline
# ═══════════════════════════════════════════════════════════════════════════

def verarbeite_datei(dateipfad: str):
    """Hauptpipeline: Extraktion → Confidence-Logik → Ablage."""
    dateiname = os.path.basename(dateipfad)

    # Bereits markierte Dateien überspringen
    if dateiname.startswith("[PRÜFEN]_"):
        return

    log.info(f"Verarbeite: {dateiname}")

    # 1. Extraktion via Claude Vision
    daten = extrahiere_rechnungsdaten(dateipfad)
    if daten is None:
        log.error(f"Extraktion fehlgeschlagen für {dateiname} – Datei bleibt in Inbox.")
        toast("Beleg nicht verarbeitet",
              f"{dateiname} - API nicht erreichbar. Datei bleibt in _Inbox, Agent später neustarten.")
        return

    confidence = daten.get("gesamt_confidence", 0)
    typ = daten.get('typ', 'Rechnung')
    za = daten.get('zahlungsart', '') or '?'
    pp = " +PayPal" if daten.get('ist_paypal') else ""
    log.info(
        f"Ergebnis: [{typ}] {daten.get('rechnungssteller')} | "
        f"{daten.get('waehrung')} {daten.get('betrag')} | "
        f"Datum: {daten.get('rechnungsdatum')} | "
        f"Zahlung: {za}{pp} | "
        f"Confidence: {confidence:.1%}"
    )

    # 2. Confidence-Logik
    if confidence >= config.CONFIDENCE_AUTO:
        log.info(f"Confidence {confidence:.1%} >= {config.CONFIDENCE_AUTO:.0%} -> Automatische Ablage")
        lege_datei_ab(dateipfad, daten)

    elif confidence >= config.CONFIDENCE_RÜCKFRAGE:
        log.info(f"Confidence {confidence:.1%} -> Bitte pruefen")
        rs = daten.get('rechnungssteller', '?')
        betrag = daten.get('betrag', 0)
        waehrung = daten.get('waehrung', '')
        toast("Beleg prüfen", f"{rs} - {waehrung} {betrag:.2f} (Confidence {confidence:.0%}) → _Inbox prüfen")
        markiere_als_pruefen(dateipfad)

    else:
        log.warning(f"Confidence {confidence:.1%} < {config.CONFIDENCE_RÜCKFRAGE:.0%} → [PRÜFEN]")
        markiere_als_pruefen(dateipfad)


# ═══════════════════════════════════════════════════════════════════════════
#  Watchdog File Watcher
# ═══════════════════════════════════════════════════════════════════════════

class BelegHandler(FileSystemEventHandler):
    """Reagiert auf neue Dateien in der Inbox."""

    def __init__(self):
        super().__init__()
        self._verarbeitet = set()

    def on_created(self, event):
        if event.is_directory:
            return

        dateipfad = event.src_path
        endung = Path(dateipfad).suffix.lower()

        if endung not in config.ERLAUBTE_ENDUNGEN:
            return

        # Bereits verarbeitete überspringen (watchdog kann doppelt feuern)
        if dateipfad in self._verarbeitet:
            return
        self._verarbeitet.add(dateipfad)

        log.info(f"Neue Datei erkannt: {os.path.basename(dateipfad)}")
        log.info(f"Warte {config.WARTE_NACH_ERKENNUNG_SEKUNDEN}s bis Datei komplett ist...")
        time.sleep(config.WARTE_NACH_ERKENNUNG_SEKUNDEN)

        # Prüfen ob Datei noch existiert (könnte verschoben worden sein)
        if not os.path.exists(dateipfad):
            log.warning(f"Datei existiert nicht mehr: {dateipfad}")
            self._verarbeitet.discard(dateipfad)
            return

        try:
            verarbeite_datei(dateipfad)
        except Exception as e:
            log.error(f"Unerwarteter Fehler bei {dateipfad}: {e}", exc_info=True)
        finally:
            self._verarbeitet.discard(dateipfad)


def verarbeite_bestehende_dateien():
    """Verarbeitet alle bereits in der Inbox liegenden Dateien beim Start."""
    if not os.path.exists(config.INBOX_PFAD):
        return

    dateien = [
        f for f in os.listdir(config.INBOX_PFAD)
        if Path(f).suffix.lower() in config.ERLAUBTE_ENDUNGEN
        and not f.startswith("[PRÜFEN]_")
    ]

    if dateien:
        log.info(f"{len(dateien)} bestehende Datei(en) in Inbox gefunden.")
        for datei in dateien:
            dateipfad = os.path.join(config.INBOX_PFAD, datei)
            try:
                verarbeite_datei(dateipfad)
            except Exception as e:
                log.error(f"Fehler bei {datei}: {e}", exc_info=True)


# ═══════════════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════
#  Erinnerungen
# ═══════════════════════════════════════════════════════════════════════════

def pruefe_erinnerungen():
    """Prueft verschiedene Bedingungen und sendet Erinnerungs-Toasts."""
    try:
        _pruefe_erinnerungen_intern()
    except Exception as e:
        log.error(f"Fehler bei Erinnerungspruefung: {e}")


def _pruefe_erinnerungen_intern():
    jetzt = datetime.now()

    # 1. Wie lange keine Belege abgelegt?
    if os.path.exists(config.EXCEL_PROTOKOLL):
        wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL, read_only=True)
        ws = wb.active
        letztes_datum = None
        unabgeglichen = 0
        for row in range(2, ws.max_row + 1):
            vd = str(ws.cell(row=row, column=12).value or "")  # Verarbeitungsdatum
            if vd:
                try:
                    d = datetime.strptime(vd[:10], "%Y-%m-%d")
                    if letztes_datum is None or d > letztes_datum:
                        letztes_datum = d
                except ValueError:
                    pass
            abgl = str(ws.cell(row=row, column=10).value or "")
            if abgl != "Ja":
                unabgeglichen += 1
        wb.close()

        if letztes_datum:
            tage_seit = (jetzt - letztes_datum).days
            if tage_seit >= 30:
                toast("Belege-Erinnerung",
                      f"Seit {tage_seit} Tagen keine Belege abgelegt! Zeit fuer einen Rundgang.")
            elif tage_seit >= 14:
                toast("Belege-Erinnerung",
                      f"Seit {tage_seit} Tagen keine neuen Belege. Hast du welche liegen?")

        # 2. Unabgeglichene Belege
        if unabgeglichen >= 5:
            toast("Abgleich fällig",
                  f"{unabgeglichen} Belege noch nicht abgeglichen. KK/Bank-Auszüge ablegen?")

    # 3. KK-Abgleich: Wie lange her?
    archiv_pfad = os.path.join(config.ABLAGE_STAMMPFAD, "_Abgleich", "archiv")
    if os.path.exists(archiv_pfad):
        kk_dateien = [f for f in os.listdir(archiv_pfad) if f.startswith("KK_")]
        if kk_dateien:
            neueste = max(os.path.getmtime(os.path.join(archiv_pfad, f)) for f in kk_dateien)
            tage_seit_kk = (jetzt.timestamp() - neueste) / 86400
            if tage_seit_kk >= 30:
                toast("KK-Abgleich fällig",
                      f"Letzter KK-Abgleich vor {int(tage_seit_kk)} Tagen. Neue Transaktionen ablegen?")

        bank_dateien = [f for f in os.listdir(archiv_pfad) if f.startswith("Bank_")]
        if bank_dateien:
            neueste = max(os.path.getmtime(os.path.join(archiv_pfad, f)) for f in bank_dateien)
            tage_seit_bank = (jetzt.timestamp() - neueste) / 86400
            if tage_seit_bank >= 30:
                toast("Bank-Abgleich fällig",
                      f"Letzter Bank-Abgleich vor {int(tage_seit_bank)} Tagen. Neue Auszüge ablegen?")

    # 4. [PRÜFEN]-Dateien in Inbox
    if os.path.exists(config.INBOX_PFAD):
        pruefen = [f for f in os.listdir(config.INBOX_PFAD) if f.startswith("[PRÜFEN]_")]
        if pruefen:
            toast("Belege prüfen",
                  f"{len(pruefen)} Beleg(e) in der Inbox warten auf manuelle Prüfung")

    # 5. Monatsanfang: Erinnerung an eBill/Dauerauftraege
    if jetzt.day <= 3:
        toast("Monatsanfang",
              f"Neuer Monat! eBill-Rechnungen und Monatsberichte prüfen?")


# ═══════════════════════════════════════════════════════════════════════════
#  Status
# ═══════════════════════════════════════════════════════════════════════════

STATUS_DATEI = os.path.join(os.path.dirname(os.path.abspath(__file__)), "beleg-agent.status")


def schreibe_status(status: str):
    """Schreibt Status in Datei damit man pruefen kann ob der Agent laeuft."""
    try:
        with open(STATUS_DATEI, "w", encoding="utf-8") as f:
            f.write(f"Status: {status}\n")
            f.write(f"Zeitpunkt: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"PID: {os.getpid()}\n")
            f.write(f"Inbox: {config.INBOX_PFAD}\n")
    except Exception:
        pass


def main():
    print()
    print("=" * 54)
    print("  BELEG-AGENT v1.0 - Rechnungsablage")
    print("=" * 54)
    print()

    # Prüfe API-Key
    if not os.environ.get("ANTHROPIC_API_KEY"):
        log.error("ANTHROPIC_API_KEY nicht gesetzt! Bitte Umgebungsvariable setzen.")
        toast("Beleg-Agent FEHLER", "ANTHROPIC_API_KEY nicht gesetzt!")
        sys.exit(1)

    # Ordnerstruktur erstellen falls noetig
    os.makedirs(config.ABLAGE_STAMMPFAD, exist_ok=True)
    os.makedirs(config.INBOX_PFAD, exist_ok=True)
    os.makedirs(config.ABGLEICH_PFAD, exist_ok=True)

    # Excel-Protokoll initialisieren
    erstelle_excel_wenn_noetig()

    # Bestehende Dateien verarbeiten
    verarbeite_bestehende_dateien()

    # Watchdog starten
    handler = BelegHandler()
    observer = Observer()
    observer.schedule(handler, config.INBOX_PFAD, recursive=False)
    observer.start()

    schreibe_status("Laeuft")
    log.info(f"Ueberwache: {config.INBOX_PFAD}")
    toast("Beleg-Agent gestartet", "Überwache Inbox auf neue Belege")

    # Erinnerungs-Timer
    letzte_erinnerung = time.time()

    try:
        while True:
            schreibe_status("Laeuft")

            # Erinnerungen alle 6 Stunden pruefen
            if time.time() - letzte_erinnerung > 6 * 3600:
                letzte_erinnerung = time.time()
                pruefe_erinnerungen()

            time.sleep(30)
    except KeyboardInterrupt:
        log.info("Beende Beleg-Agent...")
        observer.stop()

    observer.join()
    schreibe_status("Gestoppt")
    log.info("Beleg-Agent beendet.")


if __name__ == "__main__":
    main()
