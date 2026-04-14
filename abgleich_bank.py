#!/usr/bin/env python3
"""
Beleg-Agent – Bank-Abgleich
Gleicht Bankkonto-Transaktionen (CSV) mit dem Belege-Protokoll ab.
"""

import csv
import os
import re
import shutil
import sys
from datetime import datetime
from collections import defaultdict

import openpyxl

import config
import bank_profile
import offene_posten
from abgleich import _get_col, _lese_csv

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def lade_bank_transaktionen(csv_pfad: str) -> tuple[str, list[dict]]:
    """Liest Bankkonto CSV-Export anhand des konfigurierten Bank-Profils."""
    profil = bank_profile.get_profil(config.BANK_PROFIL)
    bp = profil["bank"]
    if bp is None:
        print(f"WARNUNG: Kein Bank-Profil fuer {config.BANK_PROFIL} definiert.")
        return "", []

    sp = bp["spalten"]
    zeilen = _lese_csv(csv_pfad, bp["delimiter"])

    # Waehrung erkennen
    waehrung = ""
    we = bp["waehrung_erkennung"]
    if we["methode"] == "header_zeile":
        for zeile in zeilen[:10]:
            if zeile.startswith(we["prefix"]):
                waehrung = zeile.split(we["separator"])[we["position"]].strip()
                break

    # Datenzeilen finden
    daten_start = None
    ds = bp["daten_start"]
    if ds["methode"] == "header_prefix":
        for i, zeile in enumerate(zeilen):
            if zeile.startswith(ds["prefix"]):
                daten_start = i
                break

    if daten_start is None:
        return waehrung, []

    reader = csv.DictReader(zeilen[daten_start:], delimiter=bp["delimiter"])

    # Skip-Listen aus Profil
    skip_beschreibung = bp.get("skip_beschreibung", [])
    skip_details = bp.get("skip_details", [])
    skip_buchungstext = bp.get("skip_buchungstext", [])

    transaktionen = []
    current_master_idx = None
    for row in reader:
        datum_str = _get_col(row, sp["datum"])

        # Beschreibung aus mehreren Spalten zusammenfuegen
        beschr_teile = [_get_col(row, [s]).strip().strip('"') for s in sp["beschreibung"]]
        beschreibung = " ".join(t for t in beschr_teile if t)

        # Details-Spalte
        details = _get_col(row, sp.get("details", [])).strip().strip('"')

        # Betrag bestimmen
        belastung_str = _get_col(row, sp["belastung"])
        gutschrift_str = _get_col(row, sp["gutschrift"])
        einzelbetrag_str = _get_col(row, sp.get("einzelbetrag", []))

        betrag = 0
        ist_gutschrift = False
        if gutschrift_str:
            try:
                betrag = abs(float(gutschrift_str.replace(",", ".")))
                ist_gutschrift = True
            except ValueError:
                pass
        elif belastung_str:
            try:
                betrag = abs(float(belastung_str.replace(",", ".").replace("-", "")))
            except ValueError:
                pass
        elif einzelbetrag_str:
            try:
                val = float(einzelbetrag_str.replace(",", "."))
                betrag = abs(val)
                ist_gutschrift = val > 0
            except ValueError:
                pass

        if betrag == 0:
            continue

        # Datum parsen
        if datum_str:
            try:
                datum = datetime.strptime(datum_str, bp["datum_format"]).date()
            except ValueError:
                datum = None

            # Skip-Filter aus Profil anwenden
            skip = False
            for muster in skip_beschreibung:
                if muster in beschreibung:
                    skip = True
                    break
            for muster in skip_details:
                if muster in details:
                    skip = True
                    break
            for muster in skip_buchungstext:
                if muster in beschreibung:
                    skip = True
                    break
            if skip:
                continue

            # Waehrung aus Spalte lesen (fuer Banken die es pro Zeile angeben)
            if we["methode"] == "spalte":
                zeilen_w = _get_col(row, we["spalte"])
                if zeilen_w:
                    waehrung = zeilen_w

            transaktionen.append({
                "datum": datum,
                "betrag": betrag,
                "ist_gutschrift": ist_gutschrift,
                "beschreibung": beschreibung,
                "details": details,
                "waehrung": waehrung,
                "ist_sammelauftrag_master": False,
            })
            current_master_idx = len(transaktionen) - 1
        else:
            # Unterzeile eines Sammelauftrags — markiere Vorgaenger als Master
            if current_master_idx is not None:
                transaktionen[current_master_idx]["ist_sammelauftrag_master"] = True
            transaktionen.append({
                "datum": transaktionen[-1]["datum"] if transaktionen else None,
                "betrag": betrag,
                "ist_gutschrift": ist_gutschrift,
                "beschreibung": beschreibung,
                "details": details,
                "waehrung": waehrung,
                "ist_sammelauftrag_master": False,
            })

    return waehrung, transaktionen


def erkenne_bank_typ(csv_pfad: str) -> str:
    """Erkennt Waehrung aus CSV anhand des Bank-Profils."""
    profil = bank_profile.get_profil(config.BANK_PROFIL)
    bp = profil["bank"]
    if bp is None:
        return "?"

    zeilen = _lese_csv(csv_pfad, bp["delimiter"])
    we = bp["waehrung_erkennung"]

    if we["methode"] == "header_zeile":
        for zeile in zeilen[:10]:
            if zeile.startswith(we["prefix"]):
                return zeile.split(we["separator"])[we["position"]].strip()
    elif we["methode"] == "spalte":
        # Erste Datenzeile lesen
        ds = bp["daten_start"]
        for i, zeile in enumerate(zeilen):
            if zeile.startswith(ds["prefix"]):
                reader = csv.DictReader(zeilen[i:], delimiter=bp["delimiter"])
                for row in reader:
                    w = _get_col(row, we["spalte"])
                    if w:
                        return w
                break
    return "?"


def match_bank_transaktion(trans: dict, belege: list[dict], include_matched: bool = False) -> dict | None:
    """Versucht eine Bank-Transaktion einem Beleg zuzuordnen.

    include_matched=True bezieht auch bereits als "Ja" markierte Belege ein.
    Wird im zweiten Pass genutzt, um Rolling-Export-Wiederholungen zu erkennen
    (Bank-CSVs enthalten oft Transaktionen, die schon in einem frueheren Lauf
    gematcht wurden — diese sollen nicht als neue offene Posten landen).
    """
    t_betrag = trans["betrag"]
    t_datum = trans["datum"]
    t_beschr = trans["beschreibung"].upper()
    t_details = trans["details"].upper()
    t_ist_gs = trans["ist_gutschrift"]

    beste_matches = []

    t_waehrung = trans["waehrung"].upper() if trans.get("waehrung") else ""

    for beleg in belege:
        # Bereits abgeglichene Belege ueberspringen (ausser im Recall-Pass)
        if not include_matched and beleg["abgeglichen"] == "Ja":
            continue

        b_betrag = beleg["betrag"]
        b_datum = beleg["datum"]
        b_rs = beleg["rechnungssteller"].upper()
        b_typ = beleg["typ"]
        b_waehrung = beleg["waehrung"].upper()

        # Waehrung muss stimmen (Bank-Konto-Waehrung == Beleg-Waehrung)
        if t_waehrung and b_waehrung and t_waehrung != b_waehrung:
            continue

        # Betrag muss stimmen
        if abs(t_betrag - b_betrag) > 1.0:  # Etwas mehr Toleranz fuer Bank
            continue

        # Typ-Check: Gutschrift in Bank = Gutschrift im Beleg (oder umgekehrt)
        if t_ist_gs and b_typ == "Rechnung":
            continue
        if not t_ist_gs and b_typ == "Gutschrift":
            continue

        # Datums-Toleranz: B2B-Rechnungen koennen 60-90 Tage Zahlungsfrist
        # haben, daher Fenster auf 120 Tage. Der Name-Match ist der Schutz
        # vor False Positives bei weitem Fenster.
        # Ausnahme: Dauerauftrag-Belege sind einmalig erfasst, decken aber
        # monatlich wiederkehrende Bank-Transaktionen ab — Datum ignorieren.
        if b_typ == "Dauerauftrag":
            datum_score = 0.5
        elif t_datum and b_datum:
            diff = abs((t_datum - b_datum).days)
            if diff > 120:
                continue
            datum_score = max(0, 120 - diff) / 120
        else:
            datum_score = 0.3

        # Name-Matching in Beschreibung
        name_score = 0
        rs_teile = [t for t in b_rs.split() if len(t) > 2]
        for teil in rs_teile:
            if teil in t_beschr or teil in t_details:
                name_score += 1
        if rs_teile:
            name_score = name_score / len(rs_teile)

        # Betrag-Genauigkeit
        betrag_score = 1.0 if abs(t_betrag - b_betrag) < 0.05 else 0.8

        gesamt = (name_score * 0.5) + (datum_score * 0.3) + (betrag_score * 0.2)

        if name_score > 0 or (datum_score > 0.8 and betrag_score > 0.9):
            beste_matches.append((gesamt, beleg))

    if not beste_matches:
        return None

    beste_matches.sort(key=lambda x: x[0], reverse=True)
    return beste_matches[0][1]


def main():
    print()
    print("=" * 60)
    print("  BELEG-AGENT - Bank-Abgleich")
    print(f"  Bank-Profil: {config.BANK_PROFIL}")
    print("=" * 60)
    print()

    abgleich_pfad = config.ABGLEICH_PFAD
    os.makedirs(abgleich_pfad, exist_ok=True)
    archiv_pfad = os.path.join(abgleich_pfad, "archiv")
    os.makedirs(archiv_pfad, exist_ok=True)

    # Bank-CSVs finden (nicht KK-Archiv-Dateien)
    profil = bank_profile.get_profil(config.BANK_PROFIL)
    bp = profil["bank"]
    erkennung = bp.get("erkennung", {}) if bp else {}

    csv_dateien = []
    for f in os.listdir(abgleich_pfad):
        if not f.lower().endswith(".csv"):
            continue
        pfad = os.path.join(abgleich_pfad, f)
        try:
            zeilen = _lese_csv(pfad, bp["delimiter"] if bp else ";")
            erste_zeile = zeilen[0] if zeilen else ""
            # Erkennungsmethode aus Profil
            if erkennung.get("methode") == "header_prefix":
                if erste_zeile.startswith(erkennung["prefix"]):
                    csv_dateien.append(f)
            elif erkennung.get("methode") == "content_contains":
                if erkennung["text"] in " ".join(zeilen[:5]):
                    csv_dateien.append(f)
            else:
                # Fallback: generisch pruefen
                if "Kontonummer" in erste_zeile or "IBAN" in erste_zeile:
                    csv_dateien.append(f)
        except Exception:
            pass

    if not csv_dateien:
        csv_vorhanden = [f for f in os.listdir(abgleich_pfad) if f.lower().endswith(".csv")]
        if csv_vorhanden:
            print(f"HINWEIS: {len(csv_vorhanden)} CSV-Datei(en) gefunden, aber keine als Bank-Export erkannt.")
            print(f"  Stimmt das Bank-Profil? Aktuell: '{config.BANK_PROFIL}'")
            print(f"  Aendern in config_local.py: BANK_PROFIL = \"ubs\" / \"raiffeisen\" / ...")
            print()
        print("Keine Bank-CSV-Dateien gefunden.")
        return

    print(f"Gefunden: {len(csv_dateien)} Bank-CSV-Datei(en)\n")

    # Excel laden + Struktur pruefen (mit Lock um Race Conditions zu verhindern)
    try:
        with config.excel_lock():
            wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
            ws = wb.active
            header_3 = str(ws.cell(row=1, column=3).value or "")
            if header_3 == "Betrag":
                print("FEHLER: Excel hat die alte Spaltenstruktur!")
                print("Bitte zuerst den Beleg-Agent starten — er migriert das Excel automatisch.")
                wb.close()
                return
            # Fehlende Spalten ergaenzen
            aktuelle_spalten = ws.max_column or 0
            erwartete_spalten = len(config.EXCEL_SPALTEN)
            if aktuelle_spalten < erwartete_spalten:
                for col_idx in range(aktuelle_spalten + 1, erwartete_spalten + 1):
                    ws.cell(row=1, column=col_idx).value = config.EXCEL_SPALTEN[col_idx - 1]

            belege = []
            for row_idx in range(2, ws.max_row + 1):
                datum_str = str(ws.cell(row=row_idx, column=config.COL_DATUM).value or "").strip()
                try:
                    datum = datetime.strptime(datum_str, "%Y-%m-%d").date()
                except ValueError:
                    datum = None
                try:
                    betrag = float(ws.cell(row=row_idx, column=config.COL_BETRAG).value or 0)
                except (ValueError, TypeError):
                    betrag = 0

                belege.append({
                    "row": row_idx,
                    "datum": datum,
                    "rechnungssteller": str(ws.cell(row=row_idx, column=config.COL_RECHNUNGSSTELLER).value or "").strip(),
                    "betrag": betrag,
                    "waehrung": str(ws.cell(row=row_idx, column=config.COL_WAEHRUNG).value or "").strip(),
                    "typ": str(ws.cell(row=row_idx, column=config.COL_TYP).value or "Rechnung").strip(),
                    "zahlungsart": str(ws.cell(row=row_idx, column=config.COL_ZAHLUNGSART).value or "").strip(),
                    "abgeglichen": str(ws.cell(row=row_idx, column=config.COL_ABGEGLICHEN).value or "").strip(),
                })

            ws_offen = offene_posten.ensure_sheet(wb)

            # Aufraeumen: existierende offene Posten entfernen, fuer die es inzwischen
            # einen abgeglichenen Beleg gibt (typisch bei Rolling-Exports, wo frueher
            # als offen erfasste Bank-Transaktionen nachtraeglich einen Beleg bekommen
            # haben oder bereits in einem frueheren Lauf korrekt gematcht wurden).
            cleanup_entfernt = 0
            for beleg in belege:
                if beleg["abgeglichen"] != "Ja":
                    continue
                if not beleg["betrag"]:
                    continue
                # Dauerauftrag-Belege matchen monatliche Wiederholungen — kein Datumsfilter,
                # dafuer Name+Betrag. Alle anderen Belege verwenden den normalen Datumspfad.
                if beleg["typ"] == "Dauerauftrag":
                    n = offene_posten.resolve_by_name(
                        ws_offen, beleg["rechnungssteller"], beleg["betrag"], beleg["waehrung"]
                    )
                elif beleg["datum"]:
                    n = offene_posten.resolve(
                        ws_offen, beleg["datum"], beleg["betrag"], beleg["waehrung"],
                        rechnungssteller=beleg["rechnungssteller"],
                    )
                else:
                    n = 0
                if n > 0:
                    cleanup_entfernt += n
                    print(f"  Aufgeraeumt: {beleg['rechnungssteller'][:40]} "
                          f"{beleg['waehrung']} {beleg['betrag']:.2f} "
                          f"({n} offene(r) Posten entfernt)")
            if cleanup_entfernt > 0:
                print(f"\n  -> {cleanup_entfernt} offene Posten aufgeraeumt "
                      f"(bereits durch frueheren Abgleich erfasst)\n")

            gesamt_matches = 0
            gesamt_wiederholung = 0
            gesamt_neu_za = 0
            gesamt_neu_offen = 0
            gesamt_bereits_offen = 0
            gesamt_bereits_ignoriert = 0
            ohne_beleg_liste = []

            for datei in sorted(csv_dateien):
                pfad = os.path.join(abgleich_pfad, datei)
                waehrung = erkenne_bank_typ(pfad)
                bank_typ = f"Bank {waehrung}"

                print(f"--- {bank_typ} ({datei}) ---")
                waehrung_csv, transaktionen = lade_bank_transaktionen(pfad)

                min_jahr = config.MIN_JAHR_ABGLEICH
                trans_aktuell = [t for t in transaktionen if t["datum"] and t["datum"].year >= min_jahr]
                print(f"  {len(transaktionen)} Transaktionen total, {len(trans_aktuell)} ab {min_jahr}\n")

                for trans in trans_aktuell:
                    # Sammelauftrag-Master uebergehen — die Einzelposten werden
                    # separat als Kind-Zeilen verarbeitet, die Master-Zeile ist nur
                    # ein aggregierter Summen-Eintrag ohne eigenen Beleg. Dabei auch
                    # alte offene-Posten-Eintraege fuer denselben Master loeschen
                    # (entsteht bei Rolling-Exports aus der Zeit vor dem Master-Skip).
                    if trans.get("ist_sammelauftrag_master"):
                        trans_w = trans.get("waehrung") or waehrung
                        if trans["datum"]:
                            offene_posten.resolve(
                                ws_offen, trans["datum"], trans["betrag"], trans_w
                            )
                        continue

                    match = match_bank_transaktion(trans, belege)

                    datum_str = trans["datum"].strftime("%d.%m.%Y") if trans["datum"] else "?"
                    gs = "+" if trans["ist_gutschrift"] else "-"
                    beschr_kurz = trans["beschreibung"][:50]

                    if match:
                        bereits = match["abgeglichen"] == "Ja"
                        if bereits:
                            continue

                        row = match["row"]
                        ws.cell(row=row, column=config.COL_ABGEGLICHEN).value = "Ja"

                        # Valutadatum aus Bank-Transaktion ins Protokoll schreiben
                        if trans["datum"]:
                            ws.cell(row=row, column=config.COL_VALUTADATUM).value = \
                                trans["datum"].strftime("%Y-%m-%d")

                        alte_za = ws.cell(row=row, column=config.COL_ZAHLUNGSART).value or ""
                        if not alte_za:
                            ws.cell(row=row, column=config.COL_ZAHLUNGSART).value = "Überweisung"
                            gesamt_neu_za += 1
                            za_info = " [Zahlungsart -> Überweisung]"
                        else:
                            za_info = ""

                        match["abgeglichen"] = "Ja"

                        print(f"  NEU:  {datum_str} {gs}{trans['betrag']:>10.2f} {waehrung}  {beschr_kurz}")
                        print(f"        -> {match['rechnungssteller']} ({match['waehrung']} {match['betrag']}){za_info}")
                        gesamt_matches += 1
                    else:
                        # Pass 2: koennte die Transaktion zu einem schon abgeglichenen
                        # Beleg gehoeren? Dann ist es eine Rolling-Export-Wiederholung
                        # und soll nicht erneut als offener Posten erfasst werden.
                        recall = match_bank_transaktion(trans, belege, include_matched=True)
                        if recall:
                            # Valutadatum auch nachtraeglich setzen, wenn der Beleg
                            # beim frueheren Abgleich noch keine Valuta bekommen hat
                            # und es sich nicht um einen Dauerauftrag handelt (dort
                            # wechselt die Valuta monatlich).
                            if (trans["datum"] and recall["typ"] != "Dauerauftrag"
                                    and not ws.cell(row=recall["row"], column=config.COL_VALUTADATUM).value):
                                ws.cell(row=recall["row"], column=config.COL_VALUTADATUM).value = \
                                    trans["datum"].strftime("%Y-%m-%d")
                            gesamt_wiederholung += 1
                            continue

                        trans_waehrung = trans.get("waehrung") or waehrung
                        volltext = trans["beschreibung"]
                        if trans["details"]:
                            volltext = f"{volltext} | {trans['details']}"
                        upsert_status = offene_posten.upsert(
                            ws_offen,
                            quelle=bank_typ,
                            datum=trans["datum"],
                            betrag=trans["betrag"],
                            waehrung=trans_waehrung,
                            text=volltext,
                        )
                        marker = {
                            "neu": "NEU OFFEN",
                            "bereits_offen": "BEREITS OFFEN",
                            "bereits_ignoriert": "IGNORIERT",
                        }.get(upsert_status, "KEIN BELEG")
                        print(f"  {marker}: {datum_str} {gs}{trans['betrag']:>10.2f} {trans_waehrung}  {beschr_kurz}")
                        ohne_beleg_liste.append({
                            "bank": bank_typ,
                            "datum": datum_str,
                            "betrag": trans["betrag"],
                            "gs": gs,
                            "text": beschr_kurz,
                            "upsert": upsert_status,
                        })
                        if upsert_status == "neu":
                            gesamt_neu_offen += 1
                        elif upsert_status == "bereits_offen":
                            gesamt_bereits_offen += 1
                        elif upsert_status == "bereits_ignoriert":
                            gesamt_bereits_ignoriert += 1

                print()

            # Nachtraeglicher Cleanup: die im Match-Loop frisch auf "Ja" gesetzten
            # Belege koennten noch alte offene_posten-Eintraege aus frueheren Laeufen
            # haben. Diese jetzt aufraeumen.
            nach_cleanup = 0
            for beleg in belege:
                if beleg["abgeglichen"] != "Ja":
                    continue
                if not beleg["betrag"]:
                    continue
                if beleg["typ"] == "Dauerauftrag":
                    continue  # schon im ersten Cleanup behandelt
                if not beleg["datum"]:
                    continue
                n = offene_posten.resolve(
                    ws_offen, beleg["datum"], beleg["betrag"], beleg["waehrung"],
                    rechnungssteller=beleg["rechnungssteller"],
                )
                if n > 0:
                    nach_cleanup += n
            if nach_cleanup > 0:
                print(f"Nach-Cleanup: {nach_cleanup} offene Posten entfernt\n")
                cleanup_entfernt += nach_cleanup

            # Speichern
            try:
                wb.save(config.EXCEL_PROTOKOLL)
            except PermissionError:
                print("\nFEHLER: Excel-Datei ist geöffnet! Bitte schliessen und erneut versuchen.")
                print(f"  {config.EXCEL_PROTOKOLL}")
                wb.close()
                return
            wb.close()
    except TimeoutError as e:
        print(f"\nFEHLER: {e}")
        return

    # Archivieren
    datum_str = datetime.now().strftime("%Y-%m-%d")
    for datei in csv_dateien:
        pfad = os.path.join(abgleich_pfad, datei)
        waehrung = erkenne_bank_typ(pfad)
        ziel_name = f"Bank_{waehrung}_{datum_str}.csv"
        ziel = os.path.join(archiv_pfad, ziel_name)
        shutil.move(pfad, ziel)
        print(f"Archiviert: {datei} -> archiv/{ziel_name}")

    # Zusammenfassung
    print(f"\n{'='*60}")
    print(f"  ZUSAMMENFASSUNG")
    print(f"{'='*60}")
    print(f"  Abgeglichen:           {gesamt_matches}")
    print(f"  Wiederholungen skip:   {gesamt_wiederholung}")
    if cleanup_entfernt > 0:
        print(f"  Offene Posten aufgeraeumt: {cleanup_entfernt}")
    print(f"  Zahlungsart ergaenzt:   {gesamt_neu_za}")
    print(f"  Ohne Beleg:            {len(ohne_beleg_liste)}")
    print(f"    davon neu offen:     {gesamt_neu_offen}")
    print(f"    davon schon offen:   {gesamt_bereits_offen}")
    print(f"    davon ignoriert:     {gesamt_bereits_ignoriert}")


if __name__ == "__main__":
    main()
