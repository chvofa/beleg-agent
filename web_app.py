#!/usr/bin/env python3
"""
Beleg-Agent – Web Interface
Ersetzt den System-Tray-Agent durch ein lokales Web-Dashboard.
"""

import io
import json
import os
import queue
import sys
import threading
import time
import uuid
import webbrowser
import zipfile
from datetime import datetime, date
from pathlib import Path

from flask import (
    Flask, Response, abort, jsonify, redirect, render_template,
    request, send_file, url_for,
)
from werkzeug.utils import secure_filename

# ── API-Key sicherstellen (vor beleg_agent Import) ──────────────────────────
import platform_utils
_key = platform_utils.get_api_key_from_env()
if _key:
    os.environ["ANTHROPIC_API_KEY"] = _key

import config
import openpyxl

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB


# ═══════════════════════════════════════════════════════════════════════════
#  Sicherheit
# ═══════════════════════════════════════════════════════════════════════════

def _safe_path(base_dir: str, filename: str) -> str:
    """Gibt sicheren Pfad zurück oder bricht ab (Path-Traversal-Schutz)."""
    safe_name = secure_filename(filename)
    if not safe_name:
        abort(400, "Ungültiger Dateiname")
    full = os.path.realpath(os.path.join(base_dir, safe_name))
    if not full.startswith(os.path.realpath(base_dir) + os.sep):
        abort(403, "Zugriff verweigert")
    return full


@app.before_request
def _csrf_check():
    """Prüft Origin-Header bei POST-Requests (CSRF-Schutz)."""
    if request.method != "POST":
        return
    origin = request.headers.get("Origin", "")
    referer = request.headers.get("Referer", "")
    allowed = (f"http://127.0.0.1:{WEB_PORT}", f"http://localhost:{WEB_PORT}")
    if origin:
        if not any(origin.startswith(a) for a in allowed):
            return jsonify({"error": "CSRF: Ungültiger Origin"}), 403
    elif referer:
        if not any(referer.startswith(a) for a in allowed):
            return jsonify({"error": "CSRF: Ungültiger Referer"}), 403


# ═══════════════════════════════════════════════════════════════════════════
#  SSE Event Bus
# ═══════════════════════════════════════════════════════════════════════════

class EventBus:
    """Publish-Subscribe für Server-Sent Events."""

    def __init__(self):
        self._subscribers: list[queue.Queue] = []
        self._lock = threading.Lock()

    def subscribe(self) -> queue.Queue:
        q: queue.Queue = queue.Queue(maxsize=50)
        with self._lock:
            self._subscribers.append(q)
        return q

    def unsubscribe(self, q: queue.Queue):
        with self._lock:
            self._subscribers.remove(q)

    def publish(self, event_type: str, data: dict):
        msg = {"type": event_type, "data": data}
        with self._lock:
            for q in self._subscribers:
                try:
                    q.put_nowait(msg)
                except queue.Full:
                    pass  # Client zu langsam, Event droppen


event_bus = EventBus()


# ═══════════════════════════════════════════════════════════════════════════
#  Agent Controller (ersetzt BelegTray)
# ═══════════════════════════════════════════════════════════════════════════

class AgentController:
    """Startet/stoppt den Watchdog Observer im selben Prozess."""

    def __init__(self):
        self.observer = None
        self.handler = None
        self.running = False
        self._monitor_thread = None
        self._stop_event = threading.Event()

    def start(self):
        if self.running:
            return
        import beleg_agent
        from watchdog.observers import Observer

        os.makedirs(config.ABLAGE_STAMMPFAD, exist_ok=True)
        os.makedirs(config.INBOX_PFAD, exist_ok=True)
        os.makedirs(config.ABGLEICH_PFAD, exist_ok=True)

        beleg_agent.erstelle_excel_wenn_noetig()

        self.handler = beleg_agent.BelegHandler()
        self.observer = Observer()
        self.observer.schedule(self.handler, config.INBOX_PFAD, recursive=False)
        self.observer.start()
        self.running = True

        self._stop_event.clear()
        self._monitor_thread = threading.Thread(
            target=self._monitor_loop, daemon=True
        )
        self._monitor_thread.start()

        # Bestehende Dateien verarbeiten
        threading.Thread(
            target=beleg_agent.verarbeite_bestehende_dateien, daemon=True
        ).start()

        event_bus.publish("status", {"agent": "running"})

    def stop(self):
        if not self.running:
            return
        self._stop_event.set()
        if self.observer:
            self.observer.stop()
            self.observer.join(timeout=5)
            self.observer = None
        self.running = False
        import beleg_agent
        beleg_agent.schreibe_status("Gestoppt")
        event_bus.publish("status", {"agent": "stopped"})

    def restart(self):
        self.stop()
        time.sleep(1)
        self.start()

    def get_status(self) -> dict:
        import beleg_agent
        status_datei = beleg_agent.STATUS_DATEI
        status = {
            "running": self.running,
            "status_text": "Gestoppt",
            "timestamp": None,
            "pid": os.getpid(),
            "inbox": config.INBOX_PFAD,
        }
        if os.path.exists(status_datei):
            try:
                alter = time.time() - os.path.getmtime(status_datei)
                with open(status_datei, "r") as f:
                    for line in f:
                        if line.startswith("Status:"):
                            status["status_text"] = line.split(":", 1)[1].strip()
                        elif line.startswith("Zeitpunkt:"):
                            status["timestamp"] = line.split(":", 1)[1].strip()
                status["health"] = (
                    "green" if alter < 90 else "yellow" if alter < 300 else "red"
                )
            except Exception:
                pass
        if not self.running:
            status["health"] = "red"
        return status

    def _monitor_loop(self):
        import beleg_agent
        letzte_erinnerung = time.time()
        while not self._stop_event.is_set():
            if self.running:
                beleg_agent.schreibe_status("Laeuft")
                # Health-Check
                if self.observer and not self.observer.is_alive():
                    from watchdog.observers import Observer
                    try:
                        self.observer.stop()
                        self.observer.join(timeout=5)
                    except Exception:
                        pass
                    self.observer = Observer()
                    self.observer.schedule(
                        self.handler, config.INBOX_PFAD, recursive=False
                    )
                    self.observer.start()
                    event_bus.publish("status", {"agent": "restarted"})
                    beleg_agent.verarbeite_bestehende_dateien()
                # Erinnerungen alle 6 Stunden
                if time.time() - letzte_erinnerung > 6 * 3600:
                    letzte_erinnerung = time.time()
                    beleg_agent.pruefe_erinnerungen()
            self._stop_event.wait(30)


agent = AgentController()


# ═══════════════════════════════════════════════════════════════════════════
#  Task Manager (für Abgleich/Dauerauftraege)
# ═══════════════════════════════════════════════════════════════════════════

class OutputCapture:
    """Fängt stdout ab und pusht Zeilen via SSE."""

    def __init__(self, task_id: str):
        self.task_id = task_id
        self._old_stdout = None

    def __enter__(self):
        self._old_stdout = sys.stdout
        sys.stdout = self
        return self

    def write(self, text):
        if text.strip():
            line = text.rstrip()
            task_manager.append_output(self.task_id, line)
            event_bus.publish("task_output", {
                "task_id": self.task_id, "line": line,
            })
        if self._old_stdout:
            self._old_stdout.write(text)

    def flush(self):
        if self._old_stdout:
            self._old_stdout.flush()

    def __exit__(self, *args):
        sys.stdout = self._old_stdout


class TaskManager:
    """Verwaltet Background-Tasks (Abgleich, Dauerauftraege)."""

    def __init__(self):
        self.tasks: dict[str, dict] = {}
        self._lock = threading.Lock()

    def run_task(self, task_type: str, func, *args) -> str:
        task_id = str(uuid.uuid4())[:8]
        with self._lock:
            self.tasks[task_id] = {
                "type": task_type,
                "status": "running",
                "output": [],
                "started": time.time(),
                "finished": None,
            }
        thread = threading.Thread(
            target=self._execute, args=(task_id, func, *args), daemon=True
        )
        thread.start()
        event_bus.publish("task_started", {"task_id": task_id, "type": task_type})
        return task_id

    def _execute(self, task_id, func, *args):
        with OutputCapture(task_id):
            try:
                func(*args)
                with self._lock:
                    self.tasks[task_id]["status"] = "done"
            except SystemExit:
                with self._lock:
                    self.tasks[task_id]["status"] = "error"
            except Exception as e:
                with self._lock:
                    self.tasks[task_id]["status"] = "error"
                    self.tasks[task_id]["output"].append(f"FEHLER: {e}")
        with self._lock:
            self.tasks[task_id]["finished"] = time.time()
        event_bus.publish("task_complete", {
            "task_id": task_id,
            "status": self.tasks[task_id]["status"],
        })

    def append_output(self, task_id: str, line: str):
        with self._lock:
            if task_id in self.tasks:
                self.tasks[task_id]["output"].append(line)

    def get_task(self, task_id: str) -> dict | None:
        with self._lock:
            return self.tasks.get(task_id, None)


task_manager = TaskManager()


# ═══════════════════════════════════════════════════════════════════════════
#  Helper
# ═══════════════════════════════════════════════════════════════════════════

def _resolve_ablagepfad(raw_pfad: str) -> str:
    """Löst plattformübergreifende Pfade auf (Windows→macOS via OneDrive)."""
    if not raw_pfad:
        return ""
    # Bereits lokaler Pfad?
    if os.path.sep == "/" and not raw_pfad.startswith("/"):
        # Windows-Pfad auf macOS: relativen Teil nach Stamm-Marker extrahieren
        marker = os.path.basename(config.ABLAGE_STAMMPFAD)  # z.B. "01-Belege"
        idx = raw_pfad.find(marker)
        if idx >= 0:
            rel = raw_pfad[idx + len(marker):].replace("\\", "/")
            return config.ABLAGE_STAMMPFAD + rel
    elif os.path.sep == "\\" and raw_pfad.startswith("/"):
        # macOS-Pfad auf Windows
        marker = os.path.basename(config.ABLAGE_STAMMPFAD)
        idx = raw_pfad.find(marker)
        if idx >= 0:
            rel = raw_pfad[idx + len(marker):].replace("/", "\\")
            return config.ABLAGE_STAMMPFAD + rel
    return raw_pfad


def _lade_protokoll() -> list[dict]:
    """Liest alle Zeilen aus dem Excel-Protokoll."""
    if not os.path.exists(config.EXCEL_PROTOKOLL):
        return []
    rows = []
    with config.excel_lock():
        wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            entry = {}
            for i, val in enumerate(row[: len(config.EXCEL_SPALTEN)]):
                col_name = config.EXCEL_SPALTEN[i]
                if val is None:
                    entry[col_name] = ""
                elif isinstance(val, (date, datetime)):
                    entry[col_name] = val.strftime("%Y-%m-%d")
                elif isinstance(val, float):
                    entry[col_name] = round(val, 2)
                else:
                    entry[col_name] = str(val)
            # Lokalen Pfad auflösen
            if entry.get("Ablagepfad"):
                entry["_lokaler_pfad"] = _resolve_ablagepfad(entry["Ablagepfad"])
                entry["hat_datei"] = bool(entry["_lokaler_pfad"])
            else:
                entry["_lokaler_pfad"] = ""
                entry["hat_datei"] = False
            rows.append(entry)
        wb.close()
    return rows


def _inbox_dateien() -> list[dict]:
    """Listet Dateien in der Inbox."""
    dateien = []
    if not os.path.exists(config.INBOX_PFAD):
        return dateien
    for f in sorted(os.listdir(config.INBOX_PFAD)):
        pfad = os.path.join(config.INBOX_PFAD, f)
        if not os.path.isfile(pfad):
            continue
        endung = Path(f).suffix.lower()
        if endung not in config.ERLAUBTE_ENDUNGEN:
            continue
        dateien.append({
            "name": f,
            "size": os.path.getsize(pfad),
            "modified": datetime.fromtimestamp(
                os.path.getmtime(pfad)
            ).strftime("%Y-%m-%d %H:%M"),
            "is_pruefen": f.startswith("[PRÜFEN]_") or f.startswith("[PRUEFEN]_"),
            "is_duplikat": f.startswith("[DUPLIKAT]_"),
        })
    return dateien


def _format_betrag(row: dict) -> str:
    """Formatiert Betrag mit 2 Dezimalen, Tausender-Apostroph und Währung."""
    raw = row.get("Betrag", "")
    waehrung = row.get("Währung", "")
    try:
        val = float(raw)
    except (ValueError, TypeError):
        return f"{raw} {waehrung}".strip()
    # Swiss format: 1'234.56
    if val < 0:
        prefix = "-"
        val = abs(val)
    else:
        prefix = ""
    int_part = int(val)
    dec_part = f"{val:.2f}".split(".")[1]
    # Tausender-Apostroph
    s = str(int_part)
    groups = []
    while s:
        groups.append(s[-3:])
        s = s[:-3]
    formatted = "'" .join(reversed(groups))
    result = f"{prefix}{formatted}.{dec_part}"
    if waehrung:
        result = f"{result} {waehrung}"
    return result


def _dashboard_stats() -> dict:
    """Berechnet Statistiken fürs Dashboard."""
    import offene_posten as op_mod
    protokoll = _lade_protokoll()
    inbox = _inbox_dateien()
    total = len(protokoll)
    nicht_abgeglichen = sum(
        1 for r in protokoll if r.get("Abgeglichen") != "Ja"
    )
    pruefen = sum(1 for f in inbox if f["is_pruefen"])
    duplikate = sum(1 for f in inbox if f["is_duplikat"])
    import abgleich_debitoren
    offene_posten_count = op_mod.count_offen()
    debitoren_offen = abgleich_debitoren.count_offen()

    # Formatierte Beträge für letzte Belege
    letzte = protokoll[-10:][::-1] if protokoll else []
    for b in letzte:
        b["Betrag_Formatiert"] = _format_betrag(b)

    # Zeitstempel: letzter Upload (= neuester Beleg im Protokoll)
    letzter_upload = None
    if protokoll:
        letzter_upload = protokoll[-1].get("Datum_Rechnung", "")

    # Zeitstempel: letzter Abgleich (letzter mit Abgeglichen=Ja)
    letzter_abgleich = None
    for r in reversed(protokoll):
        if r.get("Abgeglichen") == "Ja":
            letzter_abgleich = r.get("Datum_Rechnung", "")
            break

    return {
        "total_belege": total,
        "nicht_abgeglichen": nicht_abgeglichen,
        "pruefen_count": pruefen,
        "duplikat_count": duplikate,
        "inbox_count": len(inbox),
        "offene_posten_count": offene_posten_count,
        "debitoren_offen": debitoren_offen,
        "letzte_belege": letzte,
        "letzter_upload": letzter_upload,
        "letzter_abgleich": letzter_abgleich,
    }


# ═══════════════════════════════════════════════════════════════════════════
#  Seiten-Routen
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    if not os.path.exists(
        os.path.join(os.path.dirname(__file__), "config_local.py")
    ):
        return redirect(url_for("setup_page"))
    return redirect(url_for("dashboard"))


@app.route("/dashboard")
def dashboard():
    stats = _dashboard_stats()
    status = agent.get_status()
    return render_template("dashboard.html", stats=stats, status=status)


@app.route("/upload")
def upload_page():
    return render_template("upload.html")


@app.route("/protocol")
def protocol_page():
    return render_template("protocol.html")


@app.route("/reconciliation")
def reconciliation_page():
    return render_template("reconciliation.html")


@app.route("/review")
def review_page():
    return render_template("review.html")


@app.route("/offene-posten")
def offene_posten_page():
    return render_template("offene_posten.html")


@app.route("/debitoren")
def debitoren_page():
    return render_template("debitoren.html")


@app.route("/logs")
def logs_page():
    return render_template("logs.html")


@app.route("/settings")
def settings_page():
    return render_template("settings.html")


@app.route("/setup")
def setup_page():
    return render_template("setup.html")


# ═══════════════════════════════════════════════════════════════════════════
#  API-Routen
# ═══════════════════════════════════════════════════════════════════════════

# ── Status & SSE ──────────────────────────────────────────────────────────

@app.route("/api/status")
def api_status():
    return jsonify(agent.get_status())


@app.route("/api/events")
def api_events():
    q = event_bus.subscribe()

    def generate():
        try:
            while True:
                try:
                    msg = q.get(timeout=30)
                    yield (
                        f"event: {msg['type']}\n"
                        f"data: {json.dumps(msg['data'])}\n\n"
                    )
                except queue.Empty:
                    yield ": keepalive\n\n"
        finally:
            event_bus.unsubscribe(q)

    return Response(generate(), mimetype="text/event-stream")


# ── Agent Control ─────────────────────────────────────────────────────────

@app.route("/agent/start", methods=["POST"])
def agent_start():
    agent.start()
    return jsonify({"ok": True, "status": "running"})


@app.route("/agent/stop", methods=["POST"])
def agent_stop():
    agent.stop()
    return jsonify({"ok": True, "status": "stopped"})


@app.route("/agent/restart", methods=["POST"])
def agent_restart():
    agent.restart()
    return jsonify({"ok": True, "status": "running"})


# ── Upload ────────────────────────────────────────────────────────────────

@app.route("/api/upload", methods=["POST"])
def api_upload():
    if "file" not in request.files:
        return jsonify({"error": "Keine Datei"}), 400
    uploaded = request.files.getlist("file")
    gespeichert = []
    for f in uploaded:
        if not f.filename:
            continue
        safe_name = secure_filename(f.filename)
        if not safe_name:
            continue
        endung = Path(safe_name).suffix.lower()
        if endung not in config.ERLAUBTE_ENDUNGEN:
            continue
        ziel = os.path.join(config.INBOX_PFAD, safe_name)
        # Duplikat-Dateiname vermeiden
        base, ext = os.path.splitext(safe_name)
        counter = 1
        while os.path.exists(ziel):
            ziel = os.path.join(config.INBOX_PFAD, f"{base}_{counter}{ext}")
            counter += 1
        f.save(ziel)
        gespeichert.append(os.path.basename(ziel))
    event_bus.publish("upload", {"files": gespeichert})
    return jsonify({"ok": True, "files": gespeichert})


# ── Protokoll ─────────────────────────────────────────────────────────────

@app.route("/api/protocol")
def api_protocol():
    rows = _lade_protokoll()
    # Interne Pfade nicht ans Frontend senden
    for r in rows:
        r.pop("_lokaler_pfad", None)
    return jsonify(rows)


@app.route("/api/protocol/export")
def api_protocol_export():
    fmt = request.args.get("format", "xlsx")
    von = request.args.get("von", "")
    bis = request.args.get("bis", "")

    rows = _lade_protokoll()
    if von:
        rows = [r for r in rows if r.get("Datum_Rechnung", "") >= von]
    if bis:
        rows = [r for r in rows if r.get("Datum_Rechnung", "") <= bis]

    if fmt == "csv":
        output = io.StringIO()
        output.write(";".join(config.EXCEL_SPALTEN) + "\n")
        for row in rows:
            values = [str(row.get(col, "")) for col in config.EXCEL_SPALTEN]
            output.write(";".join(values) + "\n")
        buf = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        return send_file(
            buf, mimetype="text/csv",
            download_name="Belege_Export.csv", as_attachment=True,
        )

    # Excel export
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Belege"
    ws.append(config.EXCEL_SPALTEN)
    for row in rows:
        ws.append([row.get(col, "") for col in config.EXCEL_SPALTEN])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name="Belege_Export.xlsx", as_attachment=True,
    )


@app.route("/api/belege/view/<int:zeile>")
def api_beleg_view(zeile):
    """Öffnet einen einzelnen Beleg als PDF/Bild."""
    import subprocess
    rows = _lade_protokoll()
    if zeile < 0 or zeile >= len(rows):
        return jsonify({"error": "Zeile nicht gefunden"}), 404
    row = rows[zeile]
    pfad = row.get("_lokaler_pfad", "")
    if not pfad:
        return jsonify({"error": "Kein Ablagepfad vorhanden"}), 404

    # Sicherheitscheck: Pfad muss innerhalb ABLAGE_STAMMPFAD liegen
    real = os.path.realpath(pfad)
    stamm = os.path.realpath(config.ABLAGE_STAMMPFAD)
    if not real.startswith(stamm + os.sep) and real != stamm:
        return jsonify({"error": "Zugriff verweigert"}), 403

    # Datei lokal vorhanden? → direkt ausliefern
    if os.path.isfile(pfad) and os.path.getsize(pfad) > 0:
        return send_file(pfad, as_attachment=False)

    # OneDrive Files-On-Demand: Datei über System öffnen (triggert Download)
    if sys.platform == "darwin":
        subprocess.Popen(["open", pfad], stdout=subprocess.DEVNULL,
                         stderr=subprocess.DEVNULL)
        return jsonify({
            "ok": True,
            "hinweis": "Datei wird von OneDrive geladen und in Vorschau geöffnet."
        })
    elif sys.platform == "win32":
        os.startfile(pfad)
        return jsonify({"ok": True, "hinweis": "Datei wird geöffnet."})

    return jsonify({"error": "Datei nicht lokal verfügbar"}), 404


@app.route("/api/belege/download")
def api_belege_download():
    """ZIP-Download aller Belege eines Zeitraums (mit Pfad-Auflösung)."""
    von = request.args.get("von", "")
    bis = request.args.get("bis", "")

    rows = _lade_protokoll()
    if von:
        rows = [r for r in rows if r.get("Datum_Rechnung", "") >= von]
    if bis:
        rows = [r for r in rows if r.get("Datum_Rechnung", "") <= bis]

    buf = io.BytesIO()
    count = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for row in rows:
            pfad = row.get("_lokaler_pfad", "")
            if pfad and os.path.exists(pfad):
                arcname = os.path.basename(pfad)
                zf.write(pfad, arcname)
                count += 1
    buf.seek(0)
    name = f"Belege_{von or 'alle'}_{bis or 'alle'}.zip"
    return send_file(buf, mimetype="application/zip",
                     download_name=name, as_attachment=True)


# ── Abgleich ──────────────────────────────────────────────────────────────

@app.route("/api/reconciliation/upload", methods=["POST"])
def api_reconciliation_upload():
    if "file" not in request.files:
        return jsonify({"error": "Keine Datei"}), 400
    f = request.files["file"]
    if not f.filename or not f.filename.lower().endswith(".csv"):
        return jsonify({"error": "Nur CSV-Dateien"}), 400
    safe_name = secure_filename(f.filename)
    if not safe_name:
        return jsonify({"error": "Ungültiger Dateiname"}), 400
    ziel = os.path.join(config.ABGLEICH_PFAD, safe_name)
    f.save(ziel)
    return jsonify({"ok": True, "file": f.filename})


@app.route("/api/upload-csv-inbox", methods=["POST"])
def api_upload_csv_inbox():
    """Lädt eine CSV direkt in die Inbox (für Getharvest-Import)."""
    if "file" not in request.files:
        return jsonify({"error": "Keine Datei"}), 400
    f = request.files["file"]
    if not f.filename or not f.filename.lower().endswith(".csv"):
        return jsonify({"error": "Nur CSV-Dateien"}), 400
    safe_name = secure_filename(f.filename)
    if not safe_name:
        return jsonify({"error": "Ungültiger Dateiname"}), 400
    os.makedirs(config.INBOX_PFAD, exist_ok=True)
    ziel = os.path.join(config.INBOX_PFAD, safe_name)
    f.save(ziel)
    return jsonify({"ok": True, "file": safe_name})


@app.route("/api/reconciliation/kk", methods=["POST"])
def api_reconciliation_kk():
    import abgleich
    task_id = task_manager.run_task("kk_abgleich", abgleich.main)
    return jsonify({"ok": True, "task_id": task_id})


@app.route("/api/reconciliation/bank", methods=["POST"])
def api_reconciliation_bank():
    import abgleich_bank
    task_id = task_manager.run_task("bank_abgleich", abgleich_bank.main)
    return jsonify({"ok": True, "task_id": task_id})


@app.route("/api/reconciliation/dauerauftraege", methods=["POST"])
def api_reconciliation_dauerauftraege():
    import dauerauftraege
    task_id = task_manager.run_task("dauerauftraege", dauerauftraege.main)
    return jsonify({"ok": True, "task_id": task_id})


@app.route("/api/task/<task_id>")
def api_task(task_id):
    task = task_manager.get_task(task_id)
    if not task:
        return jsonify({"error": "Task nicht gefunden"}), 404
    return jsonify(task)


# ── Review ────────────────────────────────────────────────────────────────

@app.route("/api/review")
def api_review():
    return jsonify(_inbox_dateien())


@app.route("/api/review/<filename>/approve", methods=["POST"])
def api_review_approve(filename):
    import beleg_agent
    dateipfad = _safe_path(config.INBOX_PFAD, filename)
    if not os.path.exists(dateipfad):
        return jsonify({"error": "Datei nicht gefunden"}), 404

    daten = request.get_json() or {}
    # Pflichtfelder pruefen
    for feld in ["rechnungssteller", "rechnungsdatum", "betrag", "waehrung"]:
        if not daten.get(feld):
            return jsonify({"error": f"Feld '{feld}' fehlt"}), 400

    daten.setdefault("typ", "Rechnung")
    daten.setdefault("zahlungsart", "")
    daten.setdefault("ist_paypal", False)
    daten.setdefault("bemerkungen", "")
    daten["betrag"] = float(daten["betrag"])
    daten["gesamt_confidence"] = 1.0  # Manuell geprüft

    # [PRUEFEN]_ Prefix entfernen falls vorhanden
    safe_name = os.path.basename(dateipfad)
    clean_name = safe_name
    for prefix in ["[PRÜFEN]_", "[PRUEFEN]_", "[DUPLIKAT]_"]:
        if clean_name.startswith(prefix):
            clean_name = clean_name[len(prefix):]
    if clean_name != safe_name:
        neuer_pfad = os.path.join(config.INBOX_PFAD, clean_name)
        os.rename(dateipfad, neuer_pfad)
        dateipfad = neuer_pfad

    success = beleg_agent.lege_datei_ab(dateipfad, daten)
    return jsonify({"ok": success})


@app.route("/api/review/<filename>/reject", methods=["POST"])
def api_review_reject(filename):
    dateipfad = _safe_path(config.INBOX_PFAD, filename)
    if os.path.exists(dateipfad):
        os.remove(dateipfad)
    return jsonify({"ok": True})


# ── Offene Posten ─────────────────────────────────────────────────────────

@app.route("/api/offene-posten")
def api_offene_posten_list():
    import offene_posten
    posten = offene_posten.list_offen_standalone()
    return jsonify({
        "posten": posten,
        "gruende": offene_posten.GRUENDE,
    })


@app.route("/api/offene-posten/<int:row_idx>/ignorieren", methods=["POST"])
def api_offene_posten_ignorieren(row_idx):
    import offene_posten
    data = request.get_json() or {}
    grund = str(data.get("grund", "")).strip()
    notiz = str(data.get("notiz", "")).strip()
    if grund not in offene_posten.GRUENDE:
        return jsonify({"error": "Ungueltiger Grund"}), 400
    if grund == "Sonstige" and not notiz:
        return jsonify({"error": "Notiz erforderlich bei 'Sonstige'"}), 400
    ok = offene_posten.set_ignored_standalone(row_idx, grund, notiz)
    if not ok:
        return jsonify({"error": "Posten nicht gefunden oder nicht offen"}), 404
    return jsonify({"ok": True})


@app.route("/api/offene-posten/count")
def api_offene_posten_count():
    import offene_posten
    return jsonify({"count": offene_posten.count_offen()})


# ── Debitoren ─────────────────────────────────────────────────────────────

@app.route("/api/debitoren")
def api_debitoren_list():
    import abgleich_debitoren
    eintraege = abgleich_debitoren.list_aktiv_standalone()
    return jsonify({"eintraege": eintraege})


@app.route("/api/debitoren/<harvest_id>/abschreiben", methods=["POST"])
def api_debitoren_abschreiben(harvest_id):
    import abgleich_debitoren
    data = request.get_json() or {}
    notiz = str(data.get("notiz", "")).strip()
    ok = abgleich_debitoren.abschreiben_standalone(harvest_id, notiz)
    if not ok:
        return jsonify({"error": "Eintrag nicht gefunden oder nicht offen"}), 404
    return jsonify({"ok": True})


@app.route("/api/reconciliation/debitoren", methods=["POST"])
def api_reconciliation_debitoren():
    import abgleich_debitoren
    task_id = task_manager.run_task("debitoren_import", abgleich_debitoren.main)
    return jsonify({"ok": True, "task_id": task_id})


# ── Logs ──────────────────────────────────────────────────────────────────

@app.route("/api/logs")
def api_logs():
    lines = int(request.args.get("lines", 100))
    if not os.path.exists(config.LOG_DATEI):
        return jsonify({"lines": []})
    # Effizient: nur die letzten N Zeilen lesen (vom Ende der Datei)
    try:
        with open(config.LOG_DATEI, "rb") as f:
            f.seek(0, 2)
            size = f.tell()
            # Schätze ~200 Bytes pro Zeile, lese genug Puffer
            chunk = min(size, lines * 250)
            f.seek(max(0, size - chunk))
            data = f.read().decode("utf-8", errors="replace")
        result = data.splitlines()[-lines:]
    except Exception:
        result = []
    return jsonify({"lines": result})


# ── Settings ──────────────────────────────────────────────────────────────

@app.route("/api/settings")
def api_settings_get():
    return jsonify({
        "ablage_stammpfad": config.ABLAGE_STAMMPFAD,
        "bank_profil": config.BANK_PROFIL,
        "bekannte_karten": config.BEKANNTE_KARTEN,
        "confidence_auto": config.CONFIDENCE_AUTO,
        "confidence_rueckfrage": config.CONFIDENCE_RÜCKFRAGE,
        "anthropic_model": config.ANTHROPIC_MODEL,
    })


def _write_config_local(values: dict):
    """Schreibt config_local.py sicher (kein User-Input als Code)."""
    pfad = os.path.join(os.path.dirname(__file__), "config_local.py")
    lines = ['"""', "Lokale Konfiguration (generiert)", '"""', ""]
    for key, val in values.items():
        # Nur bekannte Schlüssel erlauben
        if key not in ("ABLAGE_STAMMPFAD", "BANK_PROFIL", "BEKANNTE_KARTEN",
                        "CONFIDENCE_AUTO", "CONFIDENCE_RUECKFRAGE"):
            continue
        if isinstance(val, str):
            # String: mit repr() sicher escapen (verhindert Code-Injection)
            lines.append(f"{key} = {repr(val)}")
        elif isinstance(val, (int, float)):
            lines.append(f"{key} = {val}")
        elif isinstance(val, dict):
            # Dict nur mit Strings/Zahlen als Werte (Karten-Mapping)
            safe = {repr(k): repr(v) for k, v in val.items()
                    if isinstance(k, str) and isinstance(v, str)}
            lines.append(f"{key} = {{{', '.join(f'{k}: {v}' for k, v in safe.items())}}}")
    lines.append("")
    with open(pfad, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


@app.route("/api/settings", methods=["POST"])
def api_settings_post():
    data = request.get_json() or {}
    values = {
        "ABLAGE_STAMMPFAD": str(data.get("ablage_stammpfad", config.ABLAGE_STAMMPFAD)),
        "BANK_PROFIL": str(data.get("bank_profil", config.BANK_PROFIL)),
    }
    karten = data.get("bekannte_karten", config.BEKANNTE_KARTEN)
    if karten and isinstance(karten, dict):
        values["BEKANNTE_KARTEN"] = karten
    if "confidence_auto" in data:
        values["CONFIDENCE_AUTO"] = float(data["confidence_auto"])
    if "confidence_rueckfrage" in data:
        values["CONFIDENCE_RUECKFRAGE"] = float(data["confidence_rueckfrage"])
    _write_config_local(values)
    return jsonify({"ok": True, "hinweis": "Neustart erforderlich für Änderungen"})


@app.route("/api/settings/apikey", methods=["POST"])
def api_settings_apikey():
    data = request.get_json() or {}
    key = data.get("api_key", "").strip()
    if not key:
        return jsonify({"error": "Kein API Key angegeben"}), 400
    platform_utils.set_api_key_in_env(key)
    os.environ["ANTHROPIC_API_KEY"] = key
    masked = key[:7] + "..." + key[-4:] if len(key) > 11 else key
    return jsonify({"ok": True, "masked": masked})


# ── Setup ─────────────────────────────────────────────────────────────────

@app.route("/api/setup/check")
def api_setup_check():
    """Prüft ob API-Key und config_local.py bereits existieren."""
    key = platform_utils.get_api_key_from_env()
    has_config = os.path.exists(
        os.path.join(os.path.dirname(__file__), "config_local.py")
    )
    result = {
        "has_api_key": bool(key),
        "api_key_masked": key[:7] + "..." + key[-4:] if key and len(key) > 11 else "",
        "has_config": has_config,
    }
    if has_config:
        try:
            result["ablage_stammpfad"] = config.ABLAGE_STAMMPFAD
            result["bank_profil"] = getattr(config, "BANK_PROFIL", "ubs")
        except Exception:
            pass
    return jsonify(result)


@app.route("/api/setup", methods=["POST"])
def api_setup():
    data = request.get_json() or {}

    # API Key speichern
    api_key = data.get("api_key", "")
    if api_key:
        platform_utils.set_api_key_in_env(api_key)

    # config_local.py erstellen
    stammpfad = str(data.get("ablage_stammpfad", ""))
    bank = str(data.get("bank_profil", "ubs"))
    values = {"ABLAGE_STAMMPFAD": stammpfad, "BANK_PROFIL": bank}
    karten = data.get("bekannte_karten", {})
    if karten and isinstance(karten, dict):
        values["BEKANNTE_KARTEN"] = karten
    _write_config_local(values)

    # Ordner erstellen
    stammpfad = os.path.expanduser(stammpfad)
    for sub in ["", "_Inbox", "_Abgleich", "_Dauerauftraege"]:
        os.makedirs(os.path.join(stammpfad, sub), exist_ok=True)

    return jsonify({"ok": True, "hinweis": "Setup abgeschlossen. Bitte App neustarten."})


# ═══════════════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════════════

WEB_PORT = 5001


def main():
    """Standalone-Start (ohne Tray). Fuer Tray-Start siehe tray_agent.py."""
    print()
    print("=" * 54)
    print("  BELEG-AGENT – Web Interface")
    print(f"  http://localhost:{WEB_PORT}")
    print("=" * 54)
    print()

    # Agent automatisch starten wenn konfiguriert
    if os.path.exists(os.path.join(os.path.dirname(__file__), "config_local.py")):
        try:
            agent.start()
        except Exception as e:
            print(f"  Agent-Start fehlgeschlagen: {e}")

    # Browser oeffnen
    threading.Timer(1.5, lambda: webbrowser.open(f"http://localhost:{WEB_PORT}")).start()

    app.run(host="127.0.0.1", port=WEB_PORT, threaded=True, debug=False,
            use_reloader=False)


if __name__ == "__main__":
    main()
