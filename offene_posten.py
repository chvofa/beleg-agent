#!/usr/bin/env python3
"""
Beleg-Agent – Offene Posten
Verwaltet Transaktionen (KK/Bank) fuer die bisher kein Beleg gefunden wurde.

Einziger Speicherort: Sheet "Offene_Posten" im bestehenden Belege-Excel.
Zwei Status: Offen (default) / Ignoriert (mit Pflicht-Grund).
"Erledigt" existiert nicht — ein geloester Posten wird geloescht, die Audit-Spur
liegt dann im Haupt-Protokoll (Abgeglichen=Ja).
"""

from datetime import datetime, date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

import config


SHEET_NAME = "Offene_Posten"

SPALTEN = [
    "Quelle",          # 1  KK CHF / KK EUR / Bank CHF / Bank EUR ...
    "Datum",           # 2  YYYY-MM-DD
    "Betrag",          # 3  float
    "Waehrung",        # 4
    "Buchungstext",    # 5
    "Status",          # 6  Offen / Ignoriert
    "Grund",           # 7  nur wenn Status=Ignoriert
    "Erfasst_am",      # 8  YYYY-MM-DD HH:MM (erste Erfassung)
    "Entschieden_am",  # 9  YYYY-MM-DD HH:MM (bei Ignoriert)
]

COL_QUELLE = 1
COL_DATUM = 2
COL_BETRAG = 3
COL_WAEHRUNG = 4
COL_TEXT = 5
COL_STATUS = 6
COL_GRUND = 7
COL_ERFASST = 8
COL_ENTSCHIEDEN = 9

# Gueltige Gruende fuer "Kein Beleg noetig" (Dropdown im UI)
GRUENDE = [
    "Lohn",
    "Miete",
    "Umbuchung",
    "Trinkgeld",
    "Rueckbuchung",
    "Sonstige",
]


# ═══════════════════════════════════════════════════════════════════════════
#  Sheet-Verwaltung
# ═══════════════════════════════════════════════════════════════════════════

def ensure_sheet(wb: openpyxl.Workbook):
    """Stellt sicher dass das Offene_Posten-Sheet existiert. Gibt das Sheet zurueck."""
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        # Header pruefen / ergaenzen
        for col_idx, spalte in enumerate(SPALTEN, start=1):
            if ws.cell(row=1, column=col_idx).value != spalte:
                ws.cell(row=1, column=col_idx).value = spalte
        return ws

    ws = wb.create_sheet(SHEET_NAME)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="B85450", end_color="B85450", fill_type="solid")
    for col_idx, spalte in enumerate(SPALTEN, start=1):
        zelle = ws.cell(row=1, column=col_idx, value=spalte)
        zelle.font = header_font
        zelle.fill = header_fill
        zelle.alignment = Alignment(horizontal="center")

    breiten = {1: 12, 2: 12, 3: 12, 4: 10, 5: 50, 6: 12, 7: 14, 8: 18, 9: 18}
    for i, b in breiten.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = b

    return ws


def _row_to_dict(ws, row_idx: int) -> dict:
    return {
        "row": row_idx,
        "quelle": str(ws.cell(row=row_idx, column=COL_QUELLE).value or "").strip(),
        "datum": str(ws.cell(row=row_idx, column=COL_DATUM).value or "").strip(),
        "betrag": _to_float(ws.cell(row=row_idx, column=COL_BETRAG).value),
        "waehrung": str(ws.cell(row=row_idx, column=COL_WAEHRUNG).value or "").strip(),
        "text": str(ws.cell(row=row_idx, column=COL_TEXT).value or "").strip(),
        "status": str(ws.cell(row=row_idx, column=COL_STATUS).value or "Offen").strip(),
        "grund": str(ws.cell(row=row_idx, column=COL_GRUND).value or "").strip(),
        "erfasst_am": str(ws.cell(row=row_idx, column=COL_ERFASST).value or "").strip(),
        "entschieden_am": str(ws.cell(row=row_idx, column=COL_ENTSCHIEDEN).value or "").strip(),
    }


def _to_float(val) -> float:
    if val is None or val == "":
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _norm_datum(datum) -> str:
    """Normalisiert Datum zu YYYY-MM-DD String."""
    if isinstance(datum, (date, datetime)):
        return datum.strftime("%Y-%m-%d")
    s = str(datum or "").strip()
    # Deutsches Format erkennen
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _same_posten(a_quelle: str, a_datum: str, a_betrag: float, a_waehrung: str, a_text: str,
                 b_quelle: str, b_datum: str, b_betrag: float, b_waehrung: str, b_text: str) -> bool:
    """Dedupe-Vergleich: Quelle + Datum + Betrag(0.01) + Waehrung + Buchungstext."""
    if a_quelle.strip().upper() != b_quelle.strip().upper():
        return False
    if a_datum != b_datum:
        return False
    if abs(a_betrag - b_betrag) > 0.01:
        return False
    if a_waehrung.strip().upper() != b_waehrung.strip().upper():
        return False
    if a_text.strip() != b_text.strip():
        return False
    return True


# ═══════════════════════════════════════════════════════════════════════════
#  API — wird aus Abgleich-Scripten, Upload-Flow und Web-UI aufgerufen
# ═══════════════════════════════════════════════════════════════════════════

def upsert(ws, quelle: str, datum, betrag: float, waehrung: str, text: str) -> str:
    """Fuegt einen Offenen Posten hinzu, falls er nicht schon (als Offen oder Ignoriert) existiert.

    Gibt zurueck: "neu" / "bereits_offen" / "bereits_ignoriert"
    """
    d_norm = _norm_datum(datum)
    w_norm = (waehrung or "").strip().upper()
    t_norm = (text or "").strip()

    for row_idx in range(2, ws.max_row + 1):
        r = _row_to_dict(ws, row_idx)
        if not r["quelle"]:
            continue
        if _same_posten(
            quelle, d_norm, float(betrag), w_norm, t_norm,
            r["quelle"], r["datum"], r["betrag"], r["waehrung"], r["text"],
        ):
            if r["status"] == "Ignoriert":
                return "bereits_ignoriert"
            return "bereits_offen"

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([
        quelle,
        d_norm,
        round(float(betrag), 2),
        w_norm,
        t_norm,
        "Offen",
        "",
        now,
        "",
    ])
    return "neu"


def resolve(ws, datum, betrag: float, waehrung: str,
            rechnungssteller: str = "", datum_fenster: int = 120,
            betrag_toleranz: float = 0.10) -> int:
    """Loescht offene Posten die auf einen neu abgelegten Beleg passen.

    Match-Kriterien:
    - Status muss "Offen" sein (ignorierte Posten bleiben als Audit-Spur)
    - Waehrung muss exakt stimmen
    - Betrag innerhalb betrag_toleranz (default 0.10 CHF — deckt Rundungsdifferenzen ab)
    - Datum innerhalb +/- datum_fenster Tagen (default 120 — deckt B2B-Rechnungen
      mit 60-90 Tagen Zahlungsfrist ab)
    - Falls rechnungssteller uebergeben wird: mindestens ein Namens-Wort (>2 Zeichen)
      muss im Buchungstext des offenen Postens vorkommen. Das verhindert False
      Positives bei weiterem Datumsfenster, wo z.B. zwei unabhaengige 100 CHF-
      Betraege zufaellig zusammen matchen koennten.

    Gibt die Anzahl geloeschter Zeilen zurueck.
    """
    d_norm = _norm_datum(datum)
    w_norm = (waehrung or "").strip().upper()
    try:
        d_ref = datetime.strptime(d_norm, "%Y-%m-%d").date()
    except ValueError:
        return 0

    rs_teile: list[str] = []
    if rechnungssteller:
        import re
        rs_clean = re.sub(r"\([^)]*\)", "", rechnungssteller.upper())
        rs_teile = [t for t in rs_clean.split() if len(t) > 2]

    zu_loeschen = []
    for row_idx in range(2, ws.max_row + 1):
        r = _row_to_dict(ws, row_idx)
        if r["status"] != "Offen":
            continue
        if r["waehrung"] != w_norm:
            continue
        if abs(r["betrag"] - float(betrag)) > betrag_toleranz:
            continue
        try:
            r_datum = datetime.strptime(r["datum"], "%Y-%m-%d").date()
        except ValueError:
            continue
        if abs((r_datum - d_ref).days) > datum_fenster:
            continue
        if rs_teile:
            text_upper = r["text"].upper()
            if not any(t in text_upper for t in rs_teile):
                continue
        zu_loeschen.append(row_idx)

    for row_idx in sorted(zu_loeschen, reverse=True):
        ws.delete_rows(row_idx, 1)

    return len(zu_loeschen)


def resolve_by_name(ws, rechnungssteller: str, betrag: float, waehrung: str) -> int:
    """Loescht offene Posten die zu einem Dauerauftrag-Beleg passen.

    Im Gegensatz zu resolve() wird kein Datumsfilter angewandt — Daueraufträge
    sind einmalig erfasst, decken aber monatlich wiederkehrende Bank/KK-
    Transaktionen ueber Jahre ab.

    Match-Kriterien:
    - Status muss "Offen" sein
    - Waehrung muss exakt stimmen
    - Betrag muss stimmen (Toleranz 0.02)
    - Mindestens ein Rechnungssteller-Wort > 2 Zeichen muss im Buchungstext vorkommen

    Gibt die Anzahl geloeschter Zeilen zurueck.
    """
    w_norm = (waehrung or "").strip().upper()
    rs_upper = (rechnungssteller or "").upper()
    # Klammer-Hinweise wie "(Büromiete)" ignorieren, damit der Matcher
    # greift auch wenn der OCR einen Zusatz extrahiert hat.
    import re
    rs_clean = re.sub(r"\([^)]*\)", "", rs_upper)
    rs_teile = [t for t in rs_clean.split() if len(t) > 2]
    if not rs_teile:
        return 0

    zu_loeschen = []
    for row_idx in range(2, ws.max_row + 1):
        r = _row_to_dict(ws, row_idx)
        if r["status"] != "Offen":
            continue
        if r["waehrung"] != w_norm:
            continue
        if abs(r["betrag"] - float(betrag)) > 0.02:
            continue
        text_upper = r["text"].upper()
        if not any(t in text_upper for t in rs_teile):
            continue
        zu_loeschen.append(row_idx)

    for row_idx in sorted(zu_loeschen, reverse=True):
        ws.delete_rows(row_idx, 1)

    return len(zu_loeschen)


def list_offen(ws) -> list[dict]:
    """Gibt alle offenen Posten zurueck (nur Status=Offen)."""
    result = []
    for row_idx in range(2, ws.max_row + 1):
        r = _row_to_dict(ws, row_idx)
        if not r["quelle"]:
            continue
        if r["status"] == "Offen":
            result.append(r)
    return result


def set_ignored(ws, row_idx: int, grund: str, notiz: str = "") -> bool:
    """Markiert einen offenen Posten als Ignoriert mit Pflicht-Grund.

    Gibt True zurueck wenn erfolgreich, False wenn Zeile nicht existiert oder nicht offen.
    """
    if grund not in GRUENDE:
        return False
    if row_idx < 2 or row_idx > ws.max_row:
        return False

    aktuell_status = str(ws.cell(row=row_idx, column=COL_STATUS).value or "").strip()
    if aktuell_status != "Offen":
        return False

    ws.cell(row=row_idx, column=COL_STATUS).value = "Ignoriert"
    grund_text = grund
    if grund == "Sonstige" and notiz.strip():
        grund_text = f"Sonstige: {notiz.strip()[:200]}"
    ws.cell(row=row_idx, column=COL_GRUND).value = grund_text
    ws.cell(row=row_idx, column=COL_ENTSCHIEDEN).value = datetime.now().strftime("%Y-%m-%d %H:%M")
    return True


# ═══════════════════════════════════════════════════════════════════════════
#  Standalone-Helfer mit eigenem Lock (fuer Dashboard / Upload-Hook)
# ═══════════════════════════════════════════════════════════════════════════

def count_offen() -> int:
    """Zaehlt offene Posten. Oeffnet und schliesst Excel selbst."""
    import os
    if not os.path.exists(config.EXCEL_PROTOKOLL):
        return 0
    try:
        with config.excel_lock():
            wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL, read_only=True)
            if SHEET_NAME not in wb.sheetnames:
                wb.close()
                return 0
            ws = wb[SHEET_NAME]
            anzahl = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[COL_QUELLE - 1]:
                    continue
                status = str(row[COL_STATUS - 1] or "").strip()
                if status == "Offen":
                    anzahl += 1
            wb.close()
            return anzahl
    except Exception:
        return 0


def list_offen_standalone() -> list[dict]:
    """Liest alle offenen Posten mit eigenem Lock (fuer Web-UI)."""
    import os
    if not os.path.exists(config.EXCEL_PROTOKOLL):
        return []
    with config.excel_lock():
        wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
        ensure_sheet(wb)
        ws = wb[SHEET_NAME]
        result = list_offen(ws)
        wb.close()
    return result


def resolve_standalone(datum, betrag: float, waehrung: str,
                        rechnungssteller: str = "") -> int:
    """Loest offene Posten auf (mit eigenem Lock). Aufruf aus beleg_agent.lege_datei_ab."""
    import os
    if not os.path.exists(config.EXCEL_PROTOKOLL):
        return 0
    try:
        with config.excel_lock():
            wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
            if SHEET_NAME not in wb.sheetnames:
                wb.close()
                return 0
            ws = wb[SHEET_NAME]
            anzahl = resolve(ws, datum, betrag, waehrung, rechnungssteller=rechnungssteller)
            if anzahl > 0:
                wb.save(config.EXCEL_PROTOKOLL)
            wb.close()
            return anzahl
    except Exception:
        return 0


def set_ignored_standalone(row_idx: int, grund: str, notiz: str = "") -> bool:
    """Markiert einen Posten als Ignoriert (mit eigenem Lock). Aufruf aus Web-UI."""
    import os
    if not os.path.exists(config.EXCEL_PROTOKOLL):
        return False
    try:
        with config.excel_lock():
            wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
            ensure_sheet(wb)
            ws = wb[SHEET_NAME]
            ok = set_ignored(ws, row_idx, grund, notiz)
            if ok:
                wb.save(config.EXCEL_PROTOKOLL)
            wb.close()
            return ok
    except Exception:
        return False
