#!/usr/bin/env python3
"""
Beleg-Agent – KK-Abgleich
Gleicht Kreditkarten-Transaktionen (CSV) mit dem Belege-Protokoll ab.
"""

import csv
import io
import os
import shutil
import sys
from datetime import datetime, timedelta

import openpyxl

import config
import bank_profile
import offene_posten

# Windows UTF-8
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def _get_col(row: dict, namen: list[str]) -> str:
    """Findet Spalte nach moeglichen Namen (Umlaut-Varianten)."""
    for n in namen:
        if n in row:
            return row[n].strip() if row[n] else ""
    # Fallback: fuzzy match auf Spaltennamen
    for key in row:
        for n in namen:
            if n.lower() in key.lower() or key.lower() in n.lower():
                return row[key].strip() if row[key] else ""
    return ""


def _lese_csv(csv_pfad: str, delimiter: str = ";", skip_first_line: bool = False) -> list[str]:
    """Liest CSV-Datei mit Auto-Encoding-Erkennung. Gibt Zeilen zurueck."""
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(csv_pfad, "r", encoding=enc) as f:
                inhalt = f.read()
            break
        except UnicodeDecodeError:
            continue
    else:
        with open(csv_pfad, "r", encoding="latin-1") as f:
            inhalt = f.read()

    zeilen = inhalt.strip().split("\n")
    if skip_first_line or (zeilen and zeilen[0].strip().startswith("sep=")):
        zeilen = zeilen[1:]
    return zeilen


def lade_csv_transaktionen(csv_pfad: str) -> list[dict]:
    """Liest KK CSV-Export anhand des konfigurierten Bank-Profils."""
    profil = bank_profile.get_profil(config.BANK_PROFIL)
    kk = profil["kk"]
    if kk is None:
        print(f"WARNUNG: Kein KK-Profil fuer {config.BANK_PROFIL} definiert.")
        return []

    sp = kk["spalten"]
    zeilen = _lese_csv(csv_pfad, kk["delimiter"], kk.get("skip_first_line", False))
    reader = csv.DictReader(zeilen, delimiter=kk["delimiter"])

    transaktionen = []
    for row in reader:
        try:
            betrag = float(_get_col(row, sp["betrag"]).replace(",", ".") or "0")
        except ValueError:
            betrag = 0

        datum_str = _get_col(row, sp["datum"])
        try:
            datum = datetime.strptime(datum_str, kk["datum_format"]).date()
        except ValueError:
            datum = None

        transaktionen.append({
            "datum": datum,
            "buchungstext": _get_col(row, sp["buchungstext"]),
            "betrag": betrag,
            "orig_waehrung": _get_col(row, sp["orig_waehrung"]),
            "kk_waehrung": _get_col(row, sp["kk_waehrung"]),
            "belastung": _get_col(row, sp["belastung"]),
            "gutschrift": _get_col(row, sp["gutschrift"]),
            "buchung": _get_col(row, sp.get("buchung", [])),
            "branche": _get_col(row, sp.get("branche", [])),
        })

    return transaktionen


def erkenne_kk_typ(csv_pfad: str) -> str:
    """Erkennt anhand der Waehrungsspalte ob KK CHF oder KK EUR."""
    trans = lade_csv_transaktionen(csv_pfad)
    if not trans:
        return "?"

    waehrungen = [t["kk_waehrung"] for t in trans if t["kk_waehrung"]]
    if not waehrungen:
        return "?"

    # Mehrheitsentscheid
    from collections import Counter
    haeufigste = Counter(waehrungen).most_common(1)[0][0]
    return f"KK {haeufigste}"


def pruefe_excel_struktur(wb: openpyxl.Workbook) -> None:
    """Prueft ob das Excel die aktuelle Spaltenstruktur hat. Migriert falls noetig."""
    ws = wb.active
    header_3 = str(ws.cell(row=1, column=3).value or "")
    if header_3 == "Betrag":
        print("FEHLER: Excel hat die alte Spaltenstruktur!")
        print("Bitte zuerst den Beleg-Agent starten — er migriert das Excel automatisch.")
        raise SystemExit(1)

    # Fehlende Spalten ergaenzen
    aktuelle_spalten = ws.max_column or 0
    erwartete_spalten = len(config.EXCEL_SPALTEN)
    if aktuelle_spalten < erwartete_spalten:
        for col_idx in range(aktuelle_spalten + 1, erwartete_spalten + 1):
            ws.cell(row=1, column=col_idx).value = config.EXCEL_SPALTEN[col_idx - 1]


def lade_excel_belege() -> tuple[openpyxl.Workbook, list[dict]]:
    """Laedt das Excel-Protokoll und gibt Workbook + Liste von Beleg-Dicts zurueck."""
    wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
    pruefe_excel_struktur(wb)
    ws = wb.active

    belege = []
    for row_idx in range(2, ws.max_row + 1):
        datum_str = str(ws.cell(row=row_idx, column=config.COL_DATUM).value or "").strip()
        try:
            datum = datetime.strptime(datum_str, "%Y-%m-%d").date()
        except ValueError:
            datum = None

        try:
            betrag = float(ws.cell(row=row_idx, column=config.COL_BETRAG).value or 0)
        except (ValueError, TypeError):
            betrag = 0

        belege.append({
            "row": row_idx,
            "datum": datum,
            "rechnungssteller": str(ws.cell(row=row_idx, column=config.COL_RECHNUNGSSTELLER).value or "").strip(),
            "betrag": betrag,
            "waehrung": str(ws.cell(row=row_idx, column=config.COL_WAEHRUNG).value or "").strip(),
            "typ": str(ws.cell(row=row_idx, column=config.COL_TYP).value or "").strip(),
            "zahlungsart": str(ws.cell(row=row_idx, column=config.COL_ZAHLUNGSART).value or "").strip(),
            "abgeglichen": str(ws.cell(row=row_idx, column=config.COL_ABGEGLICHEN).value or "").strip(),
        })

    return wb, belege


def match_transaktion_zu_beleg(trans: dict, belege: list[dict], kk_typ: str,
                                 include_matched: bool = False) -> dict | None:
    """Versucht eine KK-Transaktion einem Beleg zuzuordnen (fuzzy matching).

    Waehrungslogik:
    - Beleg-Waehrung muss zur Originalwaehrung der KK-Transaktion passen
    - Betrag wird gegen Originalbetrag verglichen (nicht den umgerechneten Belastungsbetrag)

    include_matched=True bezieht auch bereits als "Ja" markierte Belege ein.
    Wird im zweiten Pass genutzt, um Rolling-Export-Wiederholungen zu erkennen.
    """
    buch = trans["buchungstext"].upper()
    t_betrag = trans["betrag"]          # Originalbetrag (z.B. USD 21.62)
    t_datum = trans["datum"]
    t_orig_w = trans["orig_waehrung"].upper() if trans["orig_waehrung"] else ""

    beste_matches = []

    for beleg in belege:
        # Bereits abgeglichene Belege ueberspringen (ausser im Recall-Pass)
        if not include_matched and beleg["abgeglichen"] == "Ja":
            continue

        b_betrag = beleg["betrag"]
        b_datum = beleg["datum"]
        b_rs = beleg["rechnungssteller"].upper()
        b_waehrung = beleg["waehrung"].upper()
        b_typ = beleg["typ"]

        # 1. Waehrung muss stimmen: Beleg-Waehrung == Originalwaehrung der Transaktion
        if t_orig_w and b_waehrung and t_orig_w != b_waehrung:
            continue

        # 2. Betrag muss stimmen (kleine Toleranz fuer Rundung)
        if abs(t_betrag - b_betrag) > 0.10:
            continue

        # 3. Datums-Toleranz: KK-Einkaufsdatum kann +/- 5 Tage vom Rechnungsdatum abweichen.
        # Ausnahme: Dauerauftrag-Belege sind einmalig erfasst, decken aber monatlich
        # wiederkehrende Transaktionen ab — Datum ignorieren.
        if b_typ == "Dauerauftrag":
            datum_score = 0.5
        elif t_datum and b_datum:
            diff = abs((t_datum - b_datum).days)
            if diff > 30:
                continue
            datum_score = max(0, 30 - diff) / 30  # 1.0 bei gleichem Tag, 0.0 bei 30 Tagen
        else:
            datum_score = 0.3  # Kein Datum -> schwacher Match

        # 4. Name-Matching (Teilstring)
        name_score = 0
        rs_teile = [t for t in b_rs.split() if len(t) > 2]
        for teil in rs_teile:
            if teil in buch:
                name_score += 1
        if rs_teile:
            name_score = name_score / len(rs_teile)

        # Mindestens Namens- ODER Datumsaehnlichkeit
        gesamt_score = (name_score * 0.6) + (datum_score * 0.4)
        if name_score > 0 or (datum_score > 0.8 and abs(t_betrag - b_betrag) < 0.02):
            beste_matches.append((gesamt_score, beleg))

    if not beste_matches:
        return None

    # Bester Match
    beste_matches.sort(key=lambda x: x[0], reverse=True)
    return beste_matches[0][1]


def main():
    print()
    print("=" * 60)
    print("  BELEG-AGENT - KK-Abgleich")
    print(f"  Bank-Profil: {config.BANK_PROFIL}")
    print("=" * 60)
    print()

    if not os.path.exists(config.EXCEL_PROTOKOLL):
        print("FEHLER: Excel-Protokoll nicht gefunden!")
        return

    # 1. CSV-Dateien im Abgleich-Ordner finden
    abgleich_pfad = config.ABGLEICH_PFAD
    os.makedirs(abgleich_pfad, exist_ok=True)

    csv_dateien = [f for f in os.listdir(abgleich_pfad) if f.lower().endswith(".csv")]
    if not csv_dateien:
        print("Keine CSV-Dateien im Abgleich-Ordner gefunden.")
        return

    print(f"Gefunden: {len(csv_dateien)} CSV-Datei(en)\n")

    # 2. KK-Typ erkennen und Dateien umbenennen
    archiv_pfad = os.path.join(abgleich_pfad, "archiv")
    os.makedirs(archiv_pfad, exist_ok=True)

    kk_dateien = {}  # {"KK CHF": pfad, "KK EUR": pfad}

    for datei in csv_dateien:
        pfad = os.path.join(abgleich_pfad, datei)
        kk_typ = erkenne_kk_typ(pfad)
        print(f"  {datei} -> {kk_typ}")
        if kk_typ in ("KK CHF", "KK EUR"):
            kk_dateien[kk_typ] = pfad
        else:
            print(f"    WARNUNG: Konnte KK-Typ nicht erkennen, ueberspringe.")
            print(f"    Tipp: Stimmt das Bank-Profil? Aktuell: '{config.BANK_PROFIL}'")
            print(f"    Aendern in config_local.py: BANK_PROFIL = \"ubs\" / \"raiffeisen\" / ...")

    print()

    # 3. Excel laden (mit Lock um Race Conditions zu verhindern)
    try:
        with config.excel_lock():
            wb, belege = lade_excel_belege()
            ws = wb.active
            ws_offen = offene_posten.ensure_sheet(wb)

            # Aufraeumen: existierende offene Posten entfernen, fuer die es inzwischen
            # einen abgeglichenen Beleg gibt (Rolling-Export-Duplikate aus frueheren Laeufen).
            cleanup_entfernt = 0
            for beleg in belege:
                if beleg["abgeglichen"] != "Ja":
                    continue
                if not beleg["betrag"]:
                    continue
                if beleg["typ"] == "Dauerauftrag":
                    n = offene_posten.resolve_by_name(
                        ws_offen, beleg["rechnungssteller"], beleg["betrag"], beleg["waehrung"]
                    )
                elif beleg["datum"]:
                    n = offene_posten.resolve(
                        ws_offen, beleg["datum"], beleg["betrag"], beleg["waehrung"]
                    )
                else:
                    n = 0
                if n > 0:
                    cleanup_entfernt += n
                    print(f"  Aufgeraeumt: {beleg['rechnungssteller'][:40]} "
                          f"{beleg['waehrung']} {beleg['betrag']:.2f} "
                          f"({n} offene(r) Posten entfernt)")
            if cleanup_entfernt > 0:
                print(f"\n  -> {cleanup_entfernt} offene Posten aufgeraeumt "
                      f"(bereits durch frueheren Abgleich erfasst)\n")

            gesamt_matches = 0
            gesamt_wiederholung = 0
            gesamt_ohne_beleg = 0
            gesamt_neu_za = 0
            gesamt_neu_offen = 0
            gesamt_bereits_offen = 0
            gesamt_bereits_ignoriert = 0
            ohne_beleg_liste = []

            for kk_typ, csv_pfad in sorted(kk_dateien.items()):
                print(f"--- {kk_typ} ---")
                transaktionen = lade_csv_transaktionen(csv_pfad)

                min_jahr = config.MIN_JAHR_ABGLEICH
                trans_aktuell = [t for t in transaktionen if t["datum"] and t["datum"].year >= min_jahr]
                print(f"  {len(transaktionen)} Transaktionen total, {len(trans_aktuell)} ab {min_jahr}\n")

                for trans in trans_aktuell:
                    # Gebuehren/Zuschlaege ueberspringen
                    if not trans["buchungstext"] or "ZUSCHLAG" in trans["buchungstext"].upper():
                        continue

                    match = match_transaktion_zu_beleg(trans, belege, kk_typ)

                    datum_str = trans["datum"].strftime("%d.%m.%Y") if trans["datum"] else "?"
                    orig_w = trans["orig_waehrung"] or trans["kk_waehrung"]

                    if match:
                        row = match["row"]
                        bereits = match["abgeglichen"] == "Ja"

                        if bereits:
                            # Schon abgeglichen - still ueberspringen
                            continue

                        # Match gefunden -> Abgeglichen = Ja
                        ws.cell(row=row, column=config.COL_ABGEGLICHEN).value = "Ja"

                        # Valutadatum aus KK-Transaktion ins Protokoll schreiben
                        if trans["datum"]:
                            ws.cell(row=row, column=config.COL_VALUTADATUM).value = \
                                trans["datum"].strftime("%Y-%m-%d")

                        # Zahlungsart ergaenzen wenn leer
                        alte_za = ws.cell(row=row, column=config.COL_ZAHLUNGSART).value or ""
                        if not alte_za:
                            ws.cell(row=row, column=config.COL_ZAHLUNGSART).value = kk_typ
                            gesamt_neu_za += 1
                            za_info = f" [Zahlungsart -> {kk_typ}]"
                        else:
                            za_info = ""

                        # PayPal ergaenzen
                        if "PAYPAL" in trans["buchungstext"].upper():
                            ws.cell(row=row, column=config.COL_PAYPAL).value = "Ja"

                        # Fremdwaehrung: Belastungsbetrag + KK-Waehrung ins Excel schreiben
                        fx_info = ""
                        t_orig_w = trans["orig_waehrung"].upper() if trans["orig_waehrung"] else ""
                        t_kk_w = trans["kk_waehrung"].upper() if trans["kk_waehrung"] else ""
                        if t_orig_w and t_kk_w and t_orig_w != t_kk_w:
                            # Fremdwaehrungstransaktion
                            belastung_str = trans["belastung"]
                            if belastung_str:
                                try:
                                    belastung = float(belastung_str.replace(",", "."))
                                    ws.cell(row=row, column=config.COL_WAEHRUNG_BELASTET).value = t_kk_w
                                    ws.cell(row=row, column=config.COL_BETRAG_BELASTET).value = belastung
                                    fx_info = f" [FX: {t_kk_w} {belastung:.2f}]"
                                except ValueError:
                                    pass
                            else:
                                # Pending-Transaktion: Kurs/Belastung noch nicht bekannt
                                ws.cell(row=row, column=config.COL_WAEHRUNG_BELASTET).value = t_kk_w
                                fx_info = f" [FX: {t_kk_w} pending]"

                        # Beleg als abgeglichen markieren (in-memory auch updaten)
                        match["abgeglichen"] = "Ja"

                        print(f"  NEU:  {datum_str} {orig_w} {trans['betrag']:>10.2f}  {trans['buchungstext'][:40]}")
                        print(f"        -> {match['rechnungssteller']} ({match['waehrung']} {match['betrag']}){za_info}{fx_info}")
                        gesamt_matches += 1
                    else:
                        # Pass 2: koennte die Transaktion zu einem schon abgeglichenen
                        # Beleg gehoeren? Dann ist es eine Rolling-Export-Wiederholung.
                        recall = match_transaktion_zu_beleg(trans, belege, kk_typ, include_matched=True)
                        if recall:
                            # Valutadatum nachtraeglich setzen wenn leer
                            if (trans["datum"] and recall["typ"] != "Dauerauftrag"
                                    and not ws.cell(row=recall["row"], column=config.COL_VALUTADATUM).value):
                                ws.cell(row=recall["row"], column=config.COL_VALUTADATUM).value = \
                                    trans["datum"].strftime("%Y-%m-%d")
                            gesamt_wiederholung += 1
                            continue

                        upsert_status = offene_posten.upsert(
                            ws_offen,
                            quelle=kk_typ,
                            datum=trans["datum"],
                            betrag=trans["betrag"],
                            waehrung=orig_w,
                            text=trans["buchungstext"],
                        )
                        marker = {
                            "neu": "NEU OFFEN",
                            "bereits_offen": "BEREITS OFFEN",
                            "bereits_ignoriert": "IGNORIERT",
                        }.get(upsert_status, "KEIN BELEG")
                        print(f"  {marker}: {datum_str} {orig_w} {trans['betrag']:>10.2f}  {trans['buchungstext'][:50]}")
                        ohne_beleg_liste.append({
                            "kk": kk_typ,
                            "datum": datum_str,
                            "betrag": trans["betrag"],
                            "waehrung": orig_w,
                            "text": trans["buchungstext"][:60],
                            "upsert": upsert_status,
                        })
                        gesamt_ohne_beleg += 1
                        if upsert_status == "neu":
                            gesamt_neu_offen += 1
                        elif upsert_status == "bereits_offen":
                            gesamt_bereits_offen += 1
                        elif upsert_status == "bereits_ignoriert":
                            gesamt_bereits_ignoriert += 1

                print()

            # 4. Excel speichern
            try:
                wb.save(config.EXCEL_PROTOKOLL)
            except PermissionError:
                print("\nFEHLER: Excel-Datei ist geöffnet! Bitte schliessen und erneut versuchen.")
                print(f"  {config.EXCEL_PROTOKOLL}")
                wb.close()
                return
            wb.close()
    except TimeoutError as e:
        print(f"\nFEHLER: {e}")
        return

    # 5. CSVs archivieren
    datum_str = datetime.now().strftime("%Y-%m-%d")
    for kk_typ, csv_pfad in kk_dateien.items():
        ziel_name = f"{kk_typ.replace(' ', '_')}_{datum_str}.csv"
        ziel = os.path.join(archiv_pfad, ziel_name)
        shutil.move(csv_pfad, ziel)
        print(f"Archiviert: {os.path.basename(csv_pfad)} -> archiv/{ziel_name}")

    # 6. Zusammenfassung
    print(f"\n{'='*60}")
    print(f"  ZUSAMMENFASSUNG")
    print(f"{'='*60}")
    print(f"  Abgeglichen:              {gesamt_matches}")
    print(f"  Wiederholungen skip:      {gesamt_wiederholung}")
    if cleanup_entfernt > 0:
        print(f"  Offene Posten aufgeraeumt: {cleanup_entfernt}")
    print(f"  Zahlungsart ergaenzt:      {gesamt_neu_za}")
    print(f"  KK-Transaktionen ohne Beleg: {gesamt_ohne_beleg}")
    print(f"    davon neu offen:        {gesamt_neu_offen}")
    print(f"    davon schon offen:      {gesamt_bereits_offen}")
    print(f"    davon ignoriert:        {gesamt_bereits_ignoriert}")

    if ohne_beleg_liste:
        print(f"\n  Transaktionen OHNE passenden Beleg:")
        print(f"  {'KK':<8} {'Datum':<12} {'Betrag':>10} {'Text'}")
        print(f"  {'-'*8} {'-'*12} {'-'*10} {'-'*50}")
        for t in ohne_beleg_liste:
            print(f"  {t['kk']:<8} {t['datum']:<12} {t['waehrung']} {t['betrag']:>8.2f} {t['text']}")

    print(f"\nExcel aktualisiert: {config.EXCEL_PROTOKOLL}")


if __name__ == "__main__":
    main()
