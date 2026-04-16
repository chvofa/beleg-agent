#!/usr/bin/env python3
"""
Beleg-Agent – Debitoren-Abgleich (Getharvest)
Importiert Getharvest-Rechnungsexport (harvest_invoice_report.csv) in das Excel-Protokoll.

Ablauf:
  - Offene Rechnungen (Balance > 0) werden ins Sheet "Debitoren" eingetragen.
  - Bereits erfasste Rechnungen werden auf "Bezahlt" gesetzt, sobald
    Getharvest Balance = 0 meldet (d.h. Getharvest ist die Quelle der Wahrheit).
  - Alte bereits bezahlte Rechnungen (nicht im Sheet) werden ignoriert.

Aufruf:
    python abgleich_debitoren.py [pfad/zur/harvest_invoice_report.csv]

Ohne Argument: sucht automatisch nach harvest*.csv in _Inbox.
Nach dem Import wird die CSV nach _Abgleich/ verschoben.
"""

import csv
import os
import re
import shutil
import sys
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

import config

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


SHEET_NAME = "Debitoren"
ZAHLUNGSFRIST_TAGE = 60

STATUS_OFFEN = "Offen"
STATUS_BEZAHLT = "Bezahlt"
STATUS_ABGESCHRIEBEN = "Abgeschrieben"

SPALTEN = [
    "Harvest_ID",       # 1  Eindeutige Getharvest-Rechnungs-ID
    "Rechnungsdatum",   # 2  Issue Date (YYYY-MM-DD)
    "Faellig",          # 3  Issue Date + 60 Tage
    "Kunde",            # 4  Client
    "Betreff",          # 5  Subject
    "Betrag",           # 6  Invoice Amount (float)
    "Waehrung",         # 7  CHF / EUR / ...
    "Bezahlt_am",       # 8  Last Payment Date (leer wenn offen)
    "Status",           # 9  Offen / Bezahlt
]

COL_ID = 1
COL_DATUM = 2
COL_FAELLIG = 3
COL_KUNDE = 4
COL_BETREFF = 5
COL_BETRAG = 6
COL_WAEHRUNG = 7
COL_BEZAHLT_AM = 8
COL_STATUS = 9


# ═══════════════════════════════════════════════════════════════════════════
#  Parsing-Hilfsfunktionen
# ═══════════════════════════════════════════════════════════════════════════

def _parse_betrag(s: str) -> float:
    """Parst Getharvest-Betraege: "2'512.02" → 2512.02"""
    if not s:
        return 0.0
    cleaned = s.replace("'", "").replace(",", ".").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _parse_datum(s: str) -> str:
    """Parst Datum zu YYYY-MM-DD. Gibt '' bei leerem oder ungültigem Wert."""
    s = (s or "").strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _lese_harvest_csv(csv_pfad: str) -> list[dict]:
    """Liest Getharvest-CSV. Behandelt mehrzeilige Client-Address-Felder korrekt."""
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(csv_pfad, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            return rows
        except (UnicodeDecodeError, Exception):
            continue
    return []


# ═══════════════════════════════════════════════════════════════════════════
#  Sheet-Verwaltung
# ═══════════════════════════════════════════════════════════════════════════

def ensure_sheet(wb: openpyxl.Workbook):
    """Erstellt das Debitoren-Sheet falls nötig. Gibt ws zurück."""
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]

    ws = wb.create_sheet(SHEET_NAME)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    for col_idx, spalte in enumerate(SPALTEN, start=1):
        zelle = ws.cell(row=1, column=col_idx, value=spalte)
        zelle.font = header_font
        zelle.fill = header_fill
        zelle.alignment = Alignment(horizontal="center")

    breiten = {1: 12, 2: 14, 3: 14, 4: 32, 5: 42, 6: 12, 7: 10, 8: 14, 9: 10}
    for i, b in breiten.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = b

    return ws


def _lade_id_index(ws) -> dict[str, int]:
    """Liest bestehende Einträge. Gibt {harvest_id: row_idx} zurück."""
    index = {}
    for row_idx in range(2, ws.max_row + 1):
        raw = ws.cell(row=row_idx, column=COL_ID).value
        if raw is None:
            continue
        index[str(raw).strip()] = row_idx
    return index


# ═══════════════════════════════════════════════════════════════════════════
#  Haupt-Import
# ═══════════════════════════════════════════════════════════════════════════

def importiere_harvest(csv_pfad: str) -> dict:
    """
    Liest die Getharvest-CSV und aktualisiert das Debitoren-Sheet.

    Gibt Statistik zurück:
      neu          – neue offene Rechnungen eingetragen
      aktualisiert – bestehende Einträge von Offen → Bezahlt
      unveraendert – bereits im Sheet, kein Statuswechsel
      uebersprungen – alt bezahlt (nicht im Sheet) oder ungültige ID
    """
    if not os.path.exists(config.EXCEL_PROTOKOLL):
        return {"fehler": f"Excel nicht gefunden: {config.EXCEL_PROTOKOLL}"}

    zeilen = _lese_harvest_csv(csv_pfad)
    if not zeilen:
        return {"fehler": f"CSV konnte nicht gelesen werden: {csv_pfad}"}

    neu = 0
    aktualisiert = 0
    unveraendert = 0
    uebersprungen = 0

    with config.excel_lock():
        wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
        ws = ensure_sheet(wb)
        id_index = _lade_id_index(ws)

        for row in zeilen:
            harvest_id_str = (row.get("ID") or "").strip()
            if not harvest_id_str:
                continue
            harvest_id = harvest_id_str

            balance = _parse_betrag(row.get("Balance", "0"))
            invoice_amount = _parse_betrag(row.get("Invoice Amount", "0"))
            waehrung = (row.get("Currency Symbol") or "CHF").strip().upper()
            kunde = (row.get("Client") or "").strip()
            betreff = (row.get("Subject") or "").strip()
            rechnungsdatum = _parse_datum(row.get("Issue Date", ""))
            bezahlt_am = _parse_datum(row.get("Last Payment Date", ""))

            faellig = ""
            if rechnungsdatum:
                try:
                    faellig = (
                        datetime.strptime(rechnungsdatum, "%Y-%m-%d")
                        + timedelta(days=ZAHLUNGSFRIST_TAGE)
                    ).strftime("%Y-%m-%d")
                except ValueError:
                    pass

            ist_bezahlt = (balance <= 0.001)

            if harvest_id in id_index:
                row_idx = id_index[harvest_id]
                aktueller_status = str(
                    ws.cell(row=row_idx, column=COL_STATUS).value or ""
                ).strip()

                if aktueller_status == "Offen" and ist_bezahlt:
                    # Rechnung wurde bezahlt seit letztem Import
                    ws.cell(row=row_idx, column=COL_BEZAHLT_AM).value = bezahlt_am
                    ws.cell(row=row_idx, column=COL_STATUS).value = "Bezahlt"
                    aktualisiert += 1
                else:
                    unveraendert += 1
            else:
                if not ist_bezahlt:
                    # Neue offene Rechnung
                    ws.append([
                        str(harvest_id),
                        rechnungsdatum,
                        faellig,
                        kunde,
                        betreff,
                        round(invoice_amount, 2),
                        waehrung,
                        "",
                        "Offen",
                    ])
                    id_index[str(harvest_id)] = ws.max_row
                    neu += 1
                else:
                    # Bezahlt: nur aufnehmen wenn in letzten 12 Monaten bezahlt
                    # (Kundennamen-Datenbank für Bereinigung der Offene Posten)
                    if bezahlt_am:
                        try:
                            tage_seit_zahlung = (
                                datetime.now() -
                                datetime.strptime(bezahlt_am, "%Y-%m-%d")
                            ).days
                        except ValueError:
                            tage_seit_zahlung = 9999
                    else:
                        tage_seit_zahlung = 9999

                    if tage_seit_zahlung <= 365:
                        ws.append([
                            str(harvest_id),
                            rechnungsdatum,
                            faellig,
                            kunde,
                            betreff,
                            round(invoice_amount, 2),
                            waehrung,
                            bezahlt_am,
                            "Bezahlt",
                        ])
                        id_index[str(harvest_id)] = ws.max_row
                        neu += 1
                    else:
                        uebersprungen += 1

        wb.save(config.EXCEL_PROTOKOLL)
        wb.close()

    bereinigt = bereinige_offene_posten()

    return {
        "neu": neu,
        "aktualisiert": aktualisiert,
        "unveraendert": unveraendert,
        "uebersprungen": uebersprungen,
        "offene_posten_bereinigt": bereinigt,
    }


# ═══════════════════════════════════════════════════════════════════════════
#  Standalone-Helfer (für Web-UI)
# ═══════════════════════════════════════════════════════════════════════════

def _row_to_dict(ws, row_idx: int) -> dict:
    def _s(col): return str(ws.cell(row=row_idx, column=col).value or "").strip()
    def _f(col):
        v = ws.cell(row=row_idx, column=col).value
        try:
            return round(float(v), 2)
        except (TypeError, ValueError):
            return 0.0
    return {
        "row": row_idx,
        "harvest_id": _s(COL_ID),
        "rechnungsdatum": _s(COL_DATUM),
        "faellig": _s(COL_FAELLIG),
        "kunde": _s(COL_KUNDE),
        "betreff": _s(COL_BETREFF),
        "betrag": _f(COL_BETRAG),
        "waehrung": _s(COL_WAEHRUNG),
        "bezahlt_am": _s(COL_BEZAHLT_AM),
        "status": _s(COL_STATUS) or STATUS_OFFEN,
    }


def list_aktiv_standalone() -> list[dict]:
    """Gibt alle Debitoren-Einträge zurück die nicht Bezahlt sind (Offen + Abgeschrieben)."""
    if not os.path.exists(config.EXCEL_PROTOKOLL):
        return []
    with config.excel_lock():
        wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
        if SHEET_NAME not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[SHEET_NAME]
        result = []
        for row_idx in range(2, ws.max_row + 1):
            r = _row_to_dict(ws, row_idx)
            if not r["harvest_id"]:
                continue
            if r["status"] != STATUS_BEZAHLT:
                result.append(r)
        wb.close()
    return result


def count_offen() -> int:
    """Zählt offene Debitoren-Rechnungen."""
    if not os.path.exists(config.EXCEL_PROTOKOLL):
        return 0
    try:
        with config.excel_lock():
            wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL, read_only=True)
            if SHEET_NAME not in wb.sheetnames:
                wb.close()
                return 0
            ws = wb[SHEET_NAME]
            anzahl = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[COL_ID - 1]:
                    continue
                if str(row[COL_STATUS - 1] or "").strip() == STATUS_OFFEN:
                    anzahl += 1
            wb.close()
            return anzahl
    except Exception:
        return 0


def _kundennamen_aus_sheet(wb) -> list[str]:
    """Gibt alle Kundennamen aus dem Debitoren-Sheet zurück."""
    if SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[SHEET_NAME]
    namen = set()
    for row_idx in range(2, ws.max_row + 1):
        name = str(ws.cell(row=row_idx, column=COL_KUNDE).value or "").strip()
        if name:
            namen.add(name)
    return list(namen)


def _name_in_buchungstext(kundenname: str, buchungstext: str) -> bool:
    """True wenn mind. 2 signifikante Wörter des Kundennamens im Buchungstext vorkommen."""
    buch = buchungstext.upper()
    name_clean = re.sub(r"[^\w\s]", " ", kundenname.upper())
    ignoriert = {"GMBH", "LLC", "AG", "LTD", "INC", "SA", "SAS", "SARL"}
    woerter = [w for w in name_clean.split() if len(w) > 3 and w not in ignoriert]
    if not woerter:
        return False
    treffer = sum(1 for w in woerter if w in buch)
    return treffer >= min(2, len(woerter))


def bereinige_offene_posten() -> int:
    """
    Sucht in Offene_Posten nach Bank-Gutschriften von bekannten Debitor-Kunden
    und markiert sie als 'Ignoriert' (Debitorenzahlung).

    Wird aufgerufen nach dem Getharvest-Import und am Ende des Bank-Abgleichs.
    Gibt Anzahl bereinigter Einträge zurück.
    """
    import offene_posten as op_mod
    if not os.path.exists(config.EXCEL_PROTOKOLL):
        return 0
    try:
        with config.excel_lock():
            wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
            kundennamen = _kundennamen_aus_sheet(wb)
            if not kundennamen or op_mod.SHEET_NAME not in wb.sheetnames:
                wb.close()
                return 0

            ws_op = wb[op_mod.SHEET_NAME]
            now = datetime.now().strftime("%Y-%m-%d %H:%M")
            bereinigt = 0

            for row_idx in range(2, ws_op.max_row + 1):
                status = str(ws_op.cell(row=row_idx, column=op_mod.COL_STATUS).value or "").strip()
                if status != "Offen":
                    continue
                buchungstext = str(ws_op.cell(row=row_idx, column=op_mod.COL_TEXT).value or "")
                # Nur Bank-Gutschriften (keine KK-Ausgaben)
                if "Gutschrift" not in buchungstext:
                    continue
                # Kundenname-Match
                for name in kundennamen:
                    if _name_in_buchungstext(name, buchungstext):
                        ws_op.cell(row=row_idx, column=op_mod.COL_STATUS).value = "Ignoriert"
                        ws_op.cell(row=row_idx, column=op_mod.COL_GRUND).value = "Debitorenzahlung"
                        ws_op.cell(row=row_idx, column=op_mod.COL_ENTSCHIEDEN).value = now
                        bereinigt += 1
                        break

            if bereinigt > 0:
                wb.save(config.EXCEL_PROTOKOLL)
            wb.close()
            return bereinigt
    except Exception:
        return 0


def abschreiben_standalone(harvest_id: str, notiz: str = "") -> bool:
    """Markiert eine offene Forderung als Abgeschrieben. Gibt True bei Erfolg zurück."""
    if not harvest_id or not os.path.exists(config.EXCEL_PROTOKOLL):
        return False
    try:
        with config.excel_lock():
            wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
            if SHEET_NAME not in wb.sheetnames:
                wb.close()
                return False
            ws = wb[SHEET_NAME]
            for row_idx in range(2, ws.max_row + 1):
                if str(ws.cell(row=row_idx, column=COL_ID).value or "").strip() != harvest_id:
                    continue
                aktuell = str(ws.cell(row=row_idx, column=COL_STATUS).value or "").strip()
                if aktuell != STATUS_OFFEN:
                    wb.close()
                    return False
                ws.cell(row=row_idx, column=COL_STATUS).value = STATUS_ABGESCHRIEBEN
                # Notiz im Betreff-Feld anhängen falls angegeben
                if notiz.strip():
                    betreff = str(ws.cell(row=row_idx, column=COL_BETREFF).value or "")
                    ws.cell(row=row_idx, column=COL_BETREFF).value = (
                        f"{betreff} [Abgeschrieben: {notiz.strip()[:200]}]"
                    )
                wb.save(config.EXCEL_PROTOKOLL)
                wb.close()
                return True
            wb.close()
            return False
    except Exception:
        return False


# ═══════════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════════

def _finde_harvest_csv() -> str | None:
    """Sucht nach harvest*.csv in _Inbox."""
    inbox = config.INBOX_PFAD
    if not os.path.isdir(inbox):
        return None
    for datei in sorted(os.listdir(inbox)):
        if datei.lower().startswith("harvest") and datei.lower().endswith(".csv"):
            return os.path.join(inbox, datei)
    return None


def main():
    if len(sys.argv) > 1:
        csv_pfad = sys.argv[1]
    else:
        csv_pfad = _finde_harvest_csv()
        if not csv_pfad:
            print("FEHLER: Kein harvest*.csv in _Inbox gefunden.")
            print("Ablage: Datei in _Inbox legen oder Pfad als Argument übergeben.")
            sys.exit(1)

    if not os.path.exists(csv_pfad):
        print(f"FEHLER: Datei nicht gefunden: {csv_pfad}")
        sys.exit(1)

    print(f"Importiere: {os.path.basename(csv_pfad)}")
    ergebnis = importiere_harvest(csv_pfad)

    if "fehler" in ergebnis:
        print(f"FEHLER: {ergebnis['fehler']}")
        sys.exit(1)

    print(f"  Neu erfasst:     {ergebnis['neu']}")
    print(f"  Aktualisiert:    {ergebnis['aktualisiert']}  (Offen → Bezahlt)")
    print(f"  Unveraendert:    {ergebnis['unveraendert']}")
    print(f"  Uebersprungen:   {ergebnis['uebersprungen']}  (aelter als 12 Monate)")
    print(f"  OP bereinigt:    {ergebnis['offene_posten_bereinigt']}  (Kundenzahlungen entfernt)")

    # CSV nach _Abgleich verschieben (Audit-Spur)
    os.makedirs(config.ABGLEICH_PFAD, exist_ok=True)
    ziel = os.path.join(config.ABGLEICH_PFAD, os.path.basename(csv_pfad))
    if os.path.abspath(csv_pfad) != os.path.abspath(ziel):
        shutil.move(csv_pfad, ziel)
        print(f"  CSV verschoben: _Abgleich/{os.path.basename(csv_pfad)}")


if __name__ == "__main__":
    main()
