#!/usr/bin/env python3
"""
Beleg-Agent – Daueraufträge erfassen
Scannt PDFs in _Dauerauftraege, extrahiert Daten via Claude Vision,
benennt um und erfasst im Excel als Typ "Dauerauftrag".
"""

import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

import config
import beleg_agent

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

DAUERAUFTRAEGE_PFAD = os.path.join(config.ABLAGE_STAMMPFAD, "_Dauerauftraege")


def main():
    print()
    print("=" * 60)
    print("  BELEG-AGENT - Dauerauftraege erfassen")
    print("=" * 60)
    print()

    os.makedirs(DAUERAUFTRAEGE_PFAD, exist_ok=True)
    beleg_agent.erstelle_excel_wenn_noetig()

    # Bereits erfasste Dateien im Excel
    wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
    ws = wb.active
    bereits = set()
    for row in range(2, ws.max_row + 1):
        pfad = str(ws.cell(row=row, column=9).value or "")
        if pfad:
            bereits.add(os.path.normpath(pfad))
    wb.close()

    # Neue Dateien in _Dauerauftraege
    dateien = [
        f for f in os.listdir(DAUERAUFTRAEGE_PFAD)
        if Path(f).suffix.lower() in config.ERLAUBTE_ENDUNGEN
    ]

    if not dateien:
        print("Keine neuen Dateien in _Dauerauftraege.")
        return

    print(f"{len(dateien)} Datei(en) gefunden.\n")

    for datei in sorted(dateien):
        pfad = os.path.join(DAUERAUFTRAEGE_PFAD, datei)

        # Schon erfasst?
        if os.path.normpath(pfad) in bereits:
            print(f"  SKIP (bereits erfasst): {datei}")
            continue

        print(f"  Verarbeite: {datei}")

        daten = beleg_agent.extrahiere_rechnungsdaten(pfad)
        if daten is None:
            print(f"    FEHLER: Konnte nicht extrahieren")
            continue

        rs = daten.get("rechnungssteller", "?")
        betrag = daten.get("betrag", 0)
        waehrung = daten.get("waehrung", "CHF")
        datum = daten.get("rechnungsdatum", "")
        conf = daten.get("gesamt_confidence", 0)

        print(f"    {rs} | {waehrung} {betrag} | Datum: {datum} | Confidence: {conf:.0%}")

        # Umbenennen: Dauerauftrag [Name] - [Währung] [Betrag].pdf
        endung = Path(datei).suffix.lower()
        sauberer_name = rs.strip()
        import re
        sauberer_name = re.sub(r'[<>:"/\\|?*]', "", sauberer_name)
        neuer_name = f"Dauerauftrag {sauberer_name} - {waehrung} {betrag:.2f}{endung}"
        neuer_pfad = os.path.join(DAUERAUFTRAEGE_PFAD, neuer_name)

        if os.path.exists(neuer_pfad) and neuer_pfad != pfad:
            print(f"    Datei existiert bereits: {neuer_name}")
        elif neuer_pfad != pfad:
            os.rename(pfad, neuer_pfad)
            print(f"    Umbenannt: {neuer_name}")
            pfad = neuer_pfad

        # Ins Excel schreiben
        wb = openpyxl.load_workbook(config.EXCEL_PROTOKOLL)
        ws = wb.active

        neue_zeile = [
            datum,
            rs,
            betrag,
            waehrung,
            "Dauerauftrag",                     # Typ
            daten.get("zahlungsart", ""),
            "Ja" if daten.get("ist_paypal") else "Nein",
            datei,                               # Originaldateiname
            pfad,                                # Ablagepfad
            "Ja",                                # Abgeglichen (Daueraufträge sind per Definition ok)
            conf,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            f"Monatlicher Dauerauftrag {waehrung} {betrag:.2f}",
        ]
        ws.append(neue_zeile)
        wb.save(config.EXCEL_PROTOKOLL)
        wb.close()

        print(f"    Excel-Eintrag geschrieben (Typ: Dauerauftrag)")
        print()

    print("Fertig.")


if __name__ == "__main__":
    main()
